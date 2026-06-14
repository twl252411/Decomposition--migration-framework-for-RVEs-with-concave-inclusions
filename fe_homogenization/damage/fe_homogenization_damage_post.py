# -*- coding: utf-8 -*-
# Single-RVE Abaqus/Python post-processing script for damage homogenization.
from __future__ import print_function

import csv
import os
from pathlib import Path

from odbAccess import openOdb

try:
    import openpyxl
except Exception:
    openpyxl = None


SCRIPT_TAG = "fe_homogenization_damage_post"
ODB_PATH = Path(os.environ.get("FE_HOMOG_DAMAGE_ODB_FILE", "Job-1.odb")).expanduser().resolve()
OUTPUT_DIR = Path(os.environ.get("FE_HOMOG_DAMAGE_OUTPUT_DIR", ".")).expanduser().resolve()
OUTPUT_STEM = os.environ.get("FE_HOMOG_DAMAGE_OUTPUT_STEM", ODB_PATH.stem)
LOAD_CASE = os.environ.get("FE_HOMOG_DAMAGE_LOAD_CASE", "tension").strip().lower()
THICKNESS_UM = float(os.environ.get("FE_HOMOG_DAMAGE_THICKNESS", "1.0"))


def _pick_history_region(step):
    keys = list(step.historyRegions.keys())
    if "Node ASSEMBLY.1" in keys:
        return step.historyRegions["Node ASSEMBLY.1"]
    for key in keys:
        history_region = step.historyRegions[key]
        outputs = history_region.historyOutputs.keys()
        if "U1" in outputs and "RF1" in outputs:
            return history_region
    raise RuntimeError("No history region with U1 and RF1 found.")


def _infer_rve_lengths_um(odb):
    xvals = []
    yvals = []
    for instance_name in odb.rootAssembly.instances.keys():
        instance = odb.rootAssembly.instances[instance_name]
        for node in instance.nodes:
            coords = node.coordinates
            xvals.append(float(coords[0]))
            yvals.append(float(coords[1]))
    if len(xvals) == 0 or len(yvals) == 0:
        raise RuntimeError("Cannot infer RVE size from ODB nodes.")
    xmin = min(xvals)
    xmax = max(xvals)
    ymin = min(yvals)
    ymax = max(yvals)
    lx_um = float(xmax - xmin)
    ly_um = float(ymax - ymin)
    l_used_um = max(lx_um, ly_um)
    if l_used_um <= 0.0:
        raise RuntimeError("Inferred non-positive RVE size from ODB.")
    return l_used_um, lx_um, ly_um


def _read_curve_rows(odb_path):
    odb = openOdb(path=str(odb_path), readOnly=True)
    try:
        l_used_um, lx_um, ly_um = _infer_rve_lengths_um(odb)
        step_name = list(odb.steps.keys())[-1]
        step = odb.steps[step_name]
        history_region = _pick_history_region(step)
        u1_data = list(history_region.historyOutputs["U1"].data)
        rf1_data = list(history_region.historyOutputs["RF1"].data)
        nrow = min(len(u1_data), len(rf1_data))
        rows = []
        for index in range(nrow):
            time_value = float(u1_data[index][0])
            u1_value = float(u1_data[index][1])
            rf1_value = float(rf1_data[index][1])
            strain_value = u1_value / l_used_um
            stress_value = rf1_value / (l_used_um * THICKNESS_UM)
            if LOAD_CASE == "compression":
                strain_value = abs(strain_value)
                stress_value = abs(stress_value)
            rows.append((time_value, strain_value, stress_value, u1_value, rf1_value))
        return rows, l_used_um, lx_um, ly_um
    finally:
        odb.close()


def _peak_row(rows):
    if not rows:
        raise RuntimeError("No history rows were extracted from ODB.")
    return max(rows, key=lambda row: row[2])


def _write_curve_csv(pathname, rows, l_used_um, lx_um, ly_um):
    with open(str(pathname), "w") as fp:
        writer = csv.writer(fp)
        writer.writerow(
            [
                "time",
                "nominal_strain",
                "nominal_stress_MPa",
                "U1_control_um",
                "RF1_control_uN",
                "L_used_um",
                "Lx_um",
                "Ly_um",
            ]
        )
        for row in rows:
            writer.writerow(
                ["%.12g" % float(value) for value in row]
                + ["%.12g" % l_used_um, "%.12g" % lx_um, "%.12g" % ly_um]
            )


def _write_summary_csv(pathname, row):
    with open(str(pathname), "w") as fp:
        writer = csv.DictWriter(fp, fieldnames=list(row.keys()))
        writer.writeheader()
        writer.writerow(row)


def _write_summary_xlsx(pathname, row):
    if openpyxl is None:
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "summary"
    headers = list(row.keys())
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)
        sheet.cell(row=2, column=col_index, value=row[header])
    workbook.save(str(pathname))


def main():
    if not ODB_PATH.is_file():
        raise IOError("Input .odb file not found: %s" % ODB_PATH)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    curve_path = OUTPUT_DIR / (OUTPUT_STEM + "_stress_strain.csv")
    summary_csv_path = OUTPUT_DIR / (OUTPUT_STEM + "_summary.csv")
    summary_xlsx_path = OUTPUT_DIR / (OUTPUT_STEM + "_summary.xlsx")

    rows, l_used_um, lx_um, ly_um = _read_curve_rows(ODB_PATH)
    peak = _peak_row(rows)
    summary_row = {
        "job_name": ODB_PATH.stem,
        "load_case": LOAD_CASE,
        "odb_path": str(ODB_PATH),
        "peak_strain": float(peak[1]),
        "peak_stress_MPa": float(peak[2]),
        "L_used_um": float(l_used_um),
        "Lx_um": float(lx_um),
        "Ly_um": float(ly_um),
        "thickness_um": float(THICKNESS_UM),
    }

    _write_curve_csv(curve_path, rows, l_used_um, lx_um, ly_um)
    _write_summary_csv(summary_csv_path, summary_row)
    _write_summary_xlsx(summary_xlsx_path, summary_row)

    print("[%s] processed_odb=%s" % (SCRIPT_TAG, ODB_PATH))
    print("[%s] wrote %s" % (SCRIPT_TAG, curve_path))
    print("[%s] wrote %s" % (SCRIPT_TAG, summary_csv_path))
    if openpyxl is not None:
        print("[%s] wrote %s" % (SCRIPT_TAG, summary_xlsx_path))


if __name__ == "__main__":
    main()
