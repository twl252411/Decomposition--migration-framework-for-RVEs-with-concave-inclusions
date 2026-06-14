# Single-RVE Abaqus/Python post-processing script for elastic homogenization.
from abaqus import *
from abaqusConstants import *
from odbAccess import openOdb
from pathlib import Path
import csv
import numpy as np
import os
try:
    _REPO_ROOT = Path(__file__).resolve().parent.parent
except NameError:
    cwd = Path.cwd().resolve()
    _REPO_ROOT = cwd.parent if cwd.name == "elastic" else cwd
vendor_dir_for_import = str(_REPO_ROOT / "_vendor")
if vendor_dir_for_import not in os.sys.path:
    os.sys.path.insert(0, vendor_dir_for_import)

import openpyxl

SCRIPT_TAG = "fe_homogenization_elastic_post"
ODB_PATH = Path(
    os.environ.get("FE_HOMOG_ELASTIC_ODB_FILE", os.environ.get("FE_HOMOG_ODB_FILE", "Job-1.odb"))
).expanduser().resolve()
OUTPUT_DIR = Path(
    os.environ.get("FE_HOMOG_ELASTIC_OUTPUT_DIR", os.environ.get("HOMOG_ELASTICITY_OUTPUT_DIR", "."))
).expanduser().resolve()
OUTPUT_STEM = os.environ.get(
    "FE_HOMOG_ELASTIC_OUTPUT_STEM",
    os.environ.get("HOMOG_ELASTICITY_OUTPUT_STEM", "homogenized_elasticity_single"),
)


def _is_running_job(odb_path):
    lock_path = odb_path.with_suffix(".lck")
    if not lock_path.exists():
        return False

    sta_path = odb_path.with_suffix(".sta")
    if not sta_path.exists():
        return True

    try:
        text = sta_path.read_text(encoding="utf-8", errors="ignore").upper()
    except Exception:
        try:
            text = sta_path.read_text(errors="ignore").upper()
        except Exception:
            return True

    if "THE ANALYSIS HAS COMPLETED SUCCESSFULLY" in text or ("ABAQUS JOB" in text and "COMPLETED" in text):
        return False
    return True


def _instance_names(odb):
    names = {name.upper(): name for name in odb.rootAssembly.instances.keys()}
    required = ["PART-1-1", "PART-2-1"]
    resolved = []
    for required_name in required:
        if required_name not in names:
            raise ValueError("Instance %s not found in %s" % (required_name, odb.name))
        resolved.append(names[required_name])
    return resolved


def _volume_average_response(odb, step_name, instance_names):
    stress_row = np.zeros(3, dtype=float)
    strain_row = np.zeros(3, dtype=float)
    rve_volume = 0.0
    step = odb.steps[step_name]
    if len(step.frames) == 0:
        raise ValueError("Step %s has no output frames" % step_name)
    frame = step.frames[-1]

    for instance_name in instance_names:
        instance = odb.rootAssembly.instances[instance_name]

        stress_field = frame.fieldOutputs["S"].getSubset(region=instance, position=INTEGRATION_POINT)
        strain_field = frame.fieldOutputs["E"].getSubset(region=instance, position=INTEGRATION_POINT)
        ivol_field = frame.fieldOutputs["IVOL"].getSubset(region=instance, position=INTEGRATION_POINT)
        evol_field = frame.fieldOutputs["EVOL"].getSubset(region=instance, position=WHOLE_ELEMENT)

        for idx in range(len(stress_field.values)):
            stress_data = stress_field.values[idx].data
            strain_data = strain_field.values[idx].data
            weight = float(ivol_field.values[idx].data)
            stress_row[0] += float(stress_data[0]) * weight
            stress_row[1] += float(stress_data[1]) * weight
            stress_row[2] += float(stress_data[3]) * weight
            strain_row[0] += float(strain_data[0]) * weight
            strain_row[1] += float(strain_data[1]) * weight
            strain_row[2] += float(strain_data[3]) * weight

        for evol_value in evol_field.values:
            rve_volume += float(evol_value.data)

    if rve_volume <= 0.0:
        raise ValueError("Non-positive RVE volume in %s step %s" % (odb.name, step_name))

    stress_row /= rve_volume
    strain_row /= rve_volume
    return stress_row, strain_row, rve_volume


def _compute_case_response(odb_path):
    odb = openOdb(path=str(odb_path), readOnly=True)
    try:
        instance_names = _instance_names(odb)
        step_names = ["Step-1", "Step-2", "Step-3"]
        for step_name in step_names:
            if step_name not in odb.steps:
                raise ValueError("Step %s not found in %s" % (step_name, odb_path))

        avg_stress = np.zeros((3, 3), dtype=float)
        avg_strain = np.zeros((3, 3), dtype=float)

        for step_index, step_name in enumerate(step_names):
            stress_row, strain_row, _ = _volume_average_response(odb, step_name, instance_names)
            avg_stress[step_index, :] = stress_row
            avg_strain[step_index, :] = strain_row

        stiffness = _stiffness_from_responses(avg_stress, avg_strain)
        compliance = np.linalg.inv(stiffness)
        engineering = _engineering_constants_from_compliance(compliance)
        return {
            "stress": avg_stress,
            "strain": avg_strain,
            "compliance": compliance,
            "stiffness": stiffness,
            "engineering": engineering,
        }
    finally:
        odb.close()


def _stiffness_from_responses(avg_stress, avg_strain):
    try:
        inv_strain = np.linalg.inv(avg_strain)
    except np.linalg.LinAlgError:
        raise ValueError("Average strain response matrix is singular and stiffness cannot be computed.")
    return (inv_strain @ avg_stress).T


def _engineering_constants_from_compliance(compliance):
    return {
        "E1": 1.0 / float(compliance[0, 0]),
        "E2": 1.0 / float(compliance[1, 1]),
        "G12": 1.0 / float(compliance[2, 2]),
    }


def _write_csv(pathname, fieldnames, rows):
    with open(pathname, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(pathname, row):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "summary"
    headers = list(row.keys())
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_index, value=row.get(header))
    workbook.save(str(pathname))


def main():
    odb_path = ODB_PATH
    output_root = OUTPUT_DIR
    output_stem = OUTPUT_STEM
    if not odb_path.is_file():
        raise IOError("Input .odb file not found: %s" % odb_path)
    output_root.mkdir(parents=True, exist_ok=True)

    if _is_running_job(odb_path):
        raise RuntimeError("ODB appears to be locked by a running Abaqus job: %s" % odb_path)

    response = _compute_case_response(odb_path)
    row = {
        "relative_path": odb_path.name,
        "E1": float(response["engineering"]["E1"]),
        "E2": float(response["engineering"]["E2"]),
        "G12": float(response["engineering"]["G12"]),
    }
    case_csv = output_root / (output_stem + ".csv")
    workbook_path = output_root / (output_stem + ".xlsx")

    _write_csv(case_csv, ["relative_path", "E1", "E2", "G12"], [row])
    _write_workbook(workbook_path, row)

    print("[%s] processed_odb=%s" % (SCRIPT_TAG, odb_path))
    print("[%s] wrote %s" % (SCRIPT_TAG, case_csv))
    print("[%s] wrote %s" % (SCRIPT_TAG, workbook_path))


if __name__ == "__main__":
    main()
