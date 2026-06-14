# Single-RVE Abaqus/Python post-processing script for thermal homogenization.
from abaqusConstants import *
from odbAccess import openOdb
import numpy as np
import os
import sys

try:
    import openpyxl
except ImportError:
    bundled_site_packages = os.path.join(
        os.path.expanduser('~'),
        '.cache',
        'codex-runtimes',
        'codex-primary-runtime',
        'dependencies',
        'python',
        'Lib',
        'site-packages',
    )
    if os.path.isdir(bundled_site_packages) and bundled_site_packages not in sys.path:
        sys.path.append(bundled_site_packages)
    import openpyxl


# ---------------------------------------------------------------------------
# User parameters. Keep these consistent with fe_homogenization_thermophysical_pre.py
SCRIPT_TAG = 'fe_homogenization_thermophysical_post'
RVE_SIZE = float(os.environ.get('FE_HOMOG_THERMAL_RVE_SIZE', os.environ.get('STAGE1_RVE_SIZE', '200.0')))
DELTA_T = 1.0
OUTPUT_FILE = os.environ.get('FE_HOMOG_THERMAL_OUTPUT_XLSX', 'Homogenized_ThermoPhysical_Constants.xlsx')

CTE_JOB_PATH = os.environ.get('FE_HOMOG_THERMAL_CTE_ODB', os.environ.get('STAGE1_CTE_ODB', 'Job-1-CTE.odb'))
COND_JOB_PATH = os.environ.get('FE_HOMOG_THERMAL_ETC_ODB', os.environ.get('STAGE1_ETC_ODB', 'Job-1-ETC.odb'))


def _weighted_average_field_from_odb(odb, step_name, field_name):
    frame = odb.steps[step_name].frames[-1]
    field = frame.fieldOutputs[field_name]
    ivol = frame.fieldOutputs['IVOL']
    total = None
    volume = 0.0
    for inst_name in ('Part-1-1', 'Part-2-1'):
        inst = odb.rootAssembly.instances[inst_name]
        f_values = field.getSubset(region=inst, position=INTEGRATION_POINT).values
        v_values = ivol.getSubset(region=inst, position=INTEGRATION_POINT).values
        for val, vol in zip(f_values, v_values):
            data = np.asarray(val.data, dtype=float)
            if total is None:
                total = np.zeros(data.shape, dtype=float)
            total += data * float(vol.data)
            volume += float(vol.data)
    if volume <= 0.0:
        raise ValueError('Non-positive integration volume while averaging {}'.format(field_name))
    return total / volume, volume


def _read_rp_displacement(odb, set_name, step_name):
    rp_set = odb.rootAssembly.nodeSets[set_name]
    frame = odb.steps[step_name].frames[-1]
    values = frame.fieldOutputs['U'].getSubset(region=rp_set).values
    if not values:
        raise ValueError('No displacement output found for {}'.format(set_name))
    return np.asarray(values[0].data, dtype=float)


def postprocess_conductivity(odb_path):
    if not os.path.exists(odb_path):
        raise IOError('Cannot find {}'.format(odb_path))
    odb = openOdb(odb_path)
    try:
        hfl_x, volume_x = _weighted_average_field_from_odb(odb, 'Step-Kx', 'HFL')
        hfl_y, volume_y = _weighted_average_field_from_odb(odb, 'Step-Ky', 'HFL')
        grad = DELTA_T / RVE_SIZE
        kxx = -float(hfl_x[0]) / grad
        kyx = -float(hfl_x[1]) / grad
        kxy = -float(hfl_y[0]) / grad
        kyy = -float(hfl_y[1]) / grad
        return np.array([[kxx, kxy], [kyx, kyy]], dtype=float), hfl_x, hfl_y, 0.5 * (volume_x + volume_y)
    finally:
        odb.close()


def postprocess_cte(odb_path):
    if not os.path.exists(odb_path):
        raise IOError('Cannot find {}'.format(odb_path))
    odb = openOdb(odb_path)
    try:
        u_rp1 = _read_rp_displacement(odb, 'Set-RP-1', 'Step-CTE')
        u_rp2 = _read_rp_displacement(odb, 'Set-RP-2', 'Step-CTE')
        alpha_x = float(u_rp1[0]) / (RVE_SIZE * DELTA_T)
        alpha_y = float(u_rp2[1]) / (RVE_SIZE * DELTA_T)
        avg_s, volume = _weighted_average_field_from_odb(odb, 'Step-CTE', 'S')
        return np.array([alpha_x, alpha_y], dtype=float), u_rp1, u_rp2, avg_s, volume
    finally:
        odb.close()


def _write_matrix(sheet, row0, col0, mat):
    for i in range(mat.shape[0]):
        for j in range(mat.shape[1]):
            sheet.cell(row=row0 + i, column=col0 + j, value=float(mat[i, j]))


def main():
    k_mat, hfl_x, hfl_y, _ = postprocess_conductivity(COND_JOB_PATH)
    alpha, _, _, avg_s, _ = postprocess_cte(CTE_JOB_PATH)
    cond_row = np.array([k_mat[0, 0], k_mat[0, 1], k_mat[1, 0], k_mat[1, 1]], dtype=float)
    cte_row = np.asarray(alpha, dtype=float)
    stress_row = np.asarray(avg_s[0:3], dtype=float)
    hfl_row = np.array([hfl_x[0], hfl_x[1], hfl_y[0], hfl_y[1]], dtype=float)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'summary'
    headers = ['case', 'K11', 'K12', 'K21', 'K22', 'alpha_x', 'alpha_y', 'avg_S11_CTE', 'avg_S22_CTE', 'avg_S12_CTE']
    for j, header in enumerate(headers):
        ws.cell(row=1, column=j + 1, value=header)
    ws.cell(row=2, column=1, value='single')
    for j, value in enumerate(cond_row, start=2):
        ws.cell(row=2, column=j, value=float(value))
    for j, value in enumerate(cte_row, start=6):
        ws.cell(row=2, column=j, value=float(value))
    for j, value in enumerate(stress_row, start=8):
        ws.cell(row=2, column=j, value=float(value))

    ws_k = wb.create_sheet('conductivity_tensor')
    ws_k.cell(row=1, column=1, value='case')
    ws_k.cell(row=1, column=2, value='K matrix')
    _write_matrix(ws_k, 2, 2, np.array([[cond_row[0], cond_row[1]], [cond_row[2], cond_row[3]]], dtype=float))
    ws_k.cell(row=2, column=1, value='single')
    ws_k.cell(row=5, column=1, value='average')
    _write_matrix(ws_k, 5, 2, np.array([[cond_row[0], cond_row[1]], [cond_row[2], cond_row[3]]], dtype=float))

    ws_h = wb.create_sheet('average_heat_flux')
    for j, header in enumerate(('case', 'HFL1_Kx', 'HFL2_Kx', 'HFL1_Ky', 'HFL2_Ky')):
        ws_h.cell(row=1, column=j + 1, value=header)
    ws_h.cell(row=2, column=1, value='single')
    for j, value in enumerate(hfl_row, start=2):
        ws_h.cell(row=2, column=j, value=float(value))
    ws_h.cell(row=3, column=1, value='average')
    for j, value in enumerate(hfl_row, start=2):
        ws_h.cell(row=3, column=j, value=float(value))

    ws_a = wb.create_sheet('thermal_expansion')
    ws_a.cell(row=1, column=1, value='case')
    ws_a.cell(row=1, column=2, value='alpha_x')
    ws_a.cell(row=1, column=3, value='alpha_y')
    ws_a.cell(row=2, column=1, value='single')
    ws_a.cell(row=2, column=2, value=float(cte_row[0]))
    ws_a.cell(row=2, column=3, value=float(cte_row[1]))
    ws_a.cell(row=3, column=1, value='average')
    ws_a.cell(row=3, column=2, value=float(cte_row[0]))
    ws_a.cell(row=3, column=3, value=float(cte_row[1]))

    wb.save(OUTPUT_FILE)
    print('[%s] processed_odb=%s' % (SCRIPT_TAG, CTE_JOB_PATH))
    print('[%s] processed_odb=%s' % (SCRIPT_TAG, COND_JOB_PATH))
    print('[%s] wrote %s' % (SCRIPT_TAG, OUTPUT_FILE))


if __name__ == '__main__':
    main()
