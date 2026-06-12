# file: convex_partition_sym_ychords_to_excel.py
# -*- coding: utf-8 -*-

import math
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

PRECISION = 6

Point = Tuple[float, float]
Polygon = List[Point]


class GeometryUtils:
    @staticmethod
    def calculate_centroid(polygon: Polygon) -> Point:
        x_sum = sum(p[0] for p in polygon)
        y_sum = sum(p[1] for p in polygon)
        return (x_sum / len(polygon), y_sum / len(polygon))

    @staticmethod
    def calculate_signed_area(polygon: Polygon) -> float:
        area = 0.0
        n = len(polygon)
        for i in range(n):
            j = (i + 1) % n
            area += polygon[i][0] * polygon[j][1] - polygon[j][0] * polygon[i][1]
        return area

    @staticmethod
    def cross(o: Point, a: Point, b: Point) -> float:
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    @staticmethod
    def distance(a: Point, b: Point) -> float:
        return math.hypot(a[0] - b[0], a[1] - b[1])

    @staticmethod
    def round_coord(coord):
        if isinstance(coord, (list, tuple)):
            return tuple(round(c, PRECISION) for c in coord)
        return round(coord, PRECISION)


class PolygonUtils:
    @staticmethod
    def ensure_ccw(polygon: Polygon, epsilon: float = 1e-12) -> Polygon:
        area = GeometryUtils.calculate_signed_area(polygon)
        if abs(area) < epsilon:
            raise ValueError(f"Degenerate polygon detected (signed area ~ 0): {area:.3e}")
        return polygon if area > 0 else list(reversed(polygon))

    @staticmethod
    def is_convex(polygon: Polygon, eps: float = 1e-12) -> bool:
        poly = PolygonUtils.ensure_ccw(polygon)
        n = len(poly)
        if n < 3:
            return False
        for i in range(n):
            a = poly[i]
            b = poly[(i + 1) % n]
            c = poly[(i + 2) % n]
            if GeometryUtils.cross(a, b, c) < -eps:
                return False
        return True

    @staticmethod
    def is_point_inside_polygon(p: Point, polygon: Polygon) -> bool:
        x, y = p
        n = len(polygon)
        inside = False
        eps = 1e-12

        for i in range(n):
            j = (i + 1) % n
            xi, yi = polygon[i]
            xj, yj = polygon[j]

            cross_val = GeometryUtils.cross(polygon[i], polygon[j], p)
            if abs(cross_val) <= 1e-10:
                if (min(xi, xj) - 1e-10 <= x <= max(xi, xj) + 1e-10 and
                        min(yi, yj) - 1e-10 <= y <= max(yi, yj) + 1e-10):
                    return True

            intersects = ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi + eps) + xi)
            if intersects:
                inside = not inside

        return inside

    @staticmethod
    def is_simple_polygon(polygon: Polygon, timeout_seconds: float = 5.0) -> bool:
        import time

        start = time.time()
        poly = polygon
        n = len(poly)
        if n < 3:
            return False

        seen = set()
        for v in poly:
            rv = GeometryUtils.round_coord(v)
            if rv in seen:
                return False
            seen.add(rv)

        for i in range(n):
            if time.time() - start > timeout_seconds:
                return False
            a = poly[i]
            b = poly[(i + 1) % n]
            for j in range(i + 2, n):
                if time.time() - start > timeout_seconds:
                    return False
                c = poly[j]
                d = poly[(j + 1) % n]
                if (j + 1) % n == i:
                    continue
                if SegmentUtils.segment_intersect(a, b, c, d):
                    return False
        return True


class SegmentUtils:
    @staticmethod
    def segment_intersect(a: Point, b: Point, c: Point, d: Point, epsilon: float = 1e-10) -> bool:
        def ccw(A, B, C) -> float:
            v = (B[0] - A[0]) * (C[1] - A[1]) - (B[1] - A[1]) * (C[0] - A[0])
            return 0.0 if abs(v) <= epsilon else v

        def on_segment(p, q, r) -> bool:
            return (min(p[0], r[0]) - epsilon <= q[0] <= max(p[0], r[0]) + epsilon and
                    min(p[1], r[1]) - epsilon <= q[1] <= max(p[1], r[1]) + epsilon and
                    ccw(p, q, r) == 0.0)

        ab_c = ccw(a, b, c)
        ab_d = ccw(a, b, d)
        cd_a = ccw(c, d, a)
        cd_b = ccw(c, d, b)

        if (ab_c * ab_d < 0.0) and (cd_a * cd_b < 0.0):
            return True

        if on_segment(a, c, b) or on_segment(a, d, b) or on_segment(c, a, d) or on_segment(c, b, d):
            return True

        return False


class TriangulationUtils:
    @staticmethod
    def ear_clipping_triangulation(polygon: Polygon) -> List[Polygon]:
        poly = PolygonUtils.ensure_ccw(polygon)
        n = len(poly)
        if n < 3:
            return []
        if n == 3:
            return [poly]

        idxs = list(range(n))
        triangles: List[Polygon] = []

        def is_reflex(i0, i1, i2) -> bool:
            return GeometryUtils.cross(poly[i0], poly[i1], poly[i2]) < -1e-12

        def point_in_tri(p: Point, a: Point, b: Point, c: Point) -> bool:
            def sgn(p1, p2, p3):
                return (p1[0] - p3[0]) * (p2[1] - p3[1]) - (p2[0] - p3[0]) * (p1[1] - p3[1])

            d1 = sgn(p, a, b)
            d2 = sgn(p, b, c)
            d3 = sgn(p, c, a)
            has_neg = (d1 < -1e-12) or (d2 < -1e-12) or (d3 < -1e-12)
            has_pos = (d1 > 1e-12) or (d2 > 1e-12) or (d3 > 1e-12)
            return not (has_neg and has_pos)

        guard = 0
        while len(idxs) > 3 and guard < 100000:
            guard += 1
            ear_found = False
            m = len(idxs)
            for k in range(m):
                i_prev = idxs[(k - 1) % m]
                i_cur = idxs[k]
                i_next = idxs[(k + 1) % m]

                if is_reflex(i_prev, i_cur, i_next):
                    continue

                a, b, c = poly[i_prev], poly[i_cur], poly[i_next]
                ok = True
                for j in idxs:
                    if j in (i_prev, i_cur, i_next):
                        continue
                    if point_in_tri(poly[j], a, b, c):
                        ok = False
                        break
                if not ok:
                    continue

                triangles.append([a, b, c])
                del idxs[k]
                ear_found = True
                break

            if not ear_found:
                break

        if len(idxs) == 3:
            triangles.append([poly[idxs[0]], poly[idxs[1]], poly[idxs[2]]])

        return triangles


@dataclass(frozen=True)
class PartitionParams:
    ar_limit: float = 3.0
    area_balance_weight: float = 2.0
    area_ratio_weight: float = 0.8
    plot: bool = True
    sheet_name: str = "outer_polygon_lobular2"
    simplify_collinear_eps: float = 1e-10
    soft_fallback: bool = True
    ysym_eps: float = 1e-6


def build_cap_polygons(poly: Polygon, u0: Point, l0: Point, u1: Point, l1: Point, side: str, eps: float) -> List[Polygon]:
    poly = PolygonUtils.ensure_ccw(poly)
    n = len(poly)

    def idx_of(p: Point) -> int:
        rp = GeometryUtils.round_coord(p)
        for i, q in enumerate(poly):
            if GeometryUtils.round_coord(q) == rp:
                return i
        return min(range(n), key=lambda i: (poly[i][0] - p[0]) ** 2 + (poly[i][1] - p[1]) ** 2)

    def walk(i0: int, i1: int) -> List[Point]:
        out = [poly[i0]]
        i = i0
        while i != i1:
            i = (i + 1) % n
            out.append(poly[i])
            if len(out) > n + 5:
                break
        return out

    def chain_extreme_x(ch: List[Point]) -> float:
        xs = [p[0] for p in ch]
        return min(xs) if side == "left" else max(xs)

    caps: List[Polygon] = []

    if GeometryUtils.distance(u0, l0) <= eps:
        v = u0
        tri = [v, u1, l1]
        tri = simplify_polygon_collinear(tri, eps=1e-12)
        if len(tri) >= 3:
            tri = PolygonUtils.ensure_ccw(tri)
            if PolygonUtils.is_convex(tri):
                caps.append(tri)
            else:
                caps.extend(TriangulationUtils.ear_clipping_triangulation(tri))
        return [PolygonUtils.ensure_ccw(p) for p in caps if len(p) >= 3 and PolygonUtils.is_convex(p)]

    iu = idx_of(u0)
    il = idx_of(l0)

    ch1 = walk(iu, il)
    ch2 = walk(il, iu)
    cap_chain = ch1 if (chain_extreme_x(ch1) <= chain_extreme_x(ch2) if side == "left" else chain_extreme_x(ch1) >= chain_extreme_x(ch2)) else ch2

    cap_chain = simplify_polygon_collinear(cap_chain, eps=1e-10)
    if len(cap_chain) >= 3:
        cap = PolygonUtils.ensure_ccw(cap_chain)
        if PolygonUtils.is_convex(cap):
            caps.append(cap)
        else:
            caps.extend(TriangulationUtils.ear_clipping_triangulation(cap))

    caps = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=1e-12)) for p in caps if len(p) >= 3]
    caps = [p for p in caps if PolygonUtils.is_convex(p)]
    return caps


def convex_partition(polygon: Polygon, params: PartitionParams) -> List[Polygon]:
    poly0 = PolygonUtils.ensure_ccw(simplify_polygon_collinear(polygon, eps=params.simplify_collinear_eps))
    if not PolygonUtils.is_simple_polygon(poly0):
        raise ValueError("Input polygon is not simple (self-intersecting).")

    parts = convex_seed_by_y_symmetric_chords(poly0, params)
    parts = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=params.simplify_collinear_eps)) for p in parts]
    parts = [p for p in parts if len(p) >= 3 and PolygonUtils.is_convex(p)]

    if not parts:
        return [poly0]

    def canon_edge(a: Point, b: Point) -> Tuple[Point, Point]:
        ra = GeometryUtils.round_coord(a)
        rb = GeometryUtils.round_coord(b)
        return (ra, rb) if ra <= rb else (rb, ra)

    def build_edge_map(polys: List[Polygon]) -> Dict[Tuple[Point, Point], List[int]]:
        edge_map: Dict[Tuple[Point, Point], List[int]] = {}
        for pid, p in enumerate(polys):
            m = len(p)
            for i in range(m):
                e = canon_edge(p[i], p[(i + 1) % m])
                edge_map.setdefault(e, []).append(pid)
        return edge_map

    def merge_two_convex(poly1: Polygon, poly2: Polygon) -> Optional[Polygon]:
        eps = 10 ** (-PRECISION)

        def same_pt(u: Point, v: Point) -> bool:
            return abs(u[0] - v[0]) < eps and abs(u[1] - v[1]) < eps

        p1 = PolygonUtils.ensure_ccw(poly1)
        p2 = PolygonUtils.ensure_ccw(poly2)
        n1, n2 = len(p1), len(p2)
        if n1 < 3 or n2 < 3:
            return None

        shared = None  # (i_in_p1, j_in_p2, mode)
        # mode = "rev" means p1 has a->b and p2 has b->a (typical for CCW polygons)
        # mode = "same" means p1 has a->b and p2 also has a->b (can happen after simplification / construction)
        for i in range(n1):
            a1 = p1[i]
            b1 = p1[(i + 1) % n1]
            for j in range(n2):
                a2 = p2[j]
                b2 = p2[(j + 1) % n2]
                if same_pt(a1, b2) and same_pt(b1, a2):
                    shared = (i, j, "rev")
                    break
                if same_pt(a1, a2) and same_pt(b1, b2):
                    shared = (i, j, "same")
                    break
            if shared is not None:
                break

        if shared is None:
            return None

        i, j, mode = shared

        if mode == "rev":
            # p1 edge: a -> b, p2 edge: b -> a
            a = p1[i]
            b = p1[(i + 1) % n1]
            # In p2, j is b and j+1 is a (because edge is b->a)
            jb = j
            ja = (j + 1) % n2
        else:
            # mode == "same"
            # p1 edge: a -> b, p2 edge: a -> b
            # Convert to a "rev-like" stitching by swapping roles in p2:
            # treat p2 as if shared edge is b -> a (i.e., reverse traversal anchor)
            a = p1[i]
            b = p1[(i + 1) % n1]
            # In p2, j is a and j+1 is b (because edge is a->b)
            ja = j
            jb = (j + 1) % n2

        merged: List[Point] = []

        # Walk p1 from b to a (inclusive) along CCW boundary
        k = (i + 1) % n1  # b
        merged.append(p1[k])
        while k != i:  # until a (index i)
            k = (k + 1) % n1
            merged.append(p1[k])

        # Now merged ends at a. Walk p2 from a to b (exclusive of a and b duplicates),
        # i.e., start after a and stop before b.
        k = (ja + 1) % n2
        while k != jb:
            merged.append(p2[k])
            k = (k + 1) % n2

        # Cleanup duplicate consecutive points
        cleaned: List[Point] = []
        for pt in merged:
            if not cleaned:
                cleaned.append(pt)
            else:
                if abs(cleaned[-1][0] - pt[0]) > eps or abs(cleaned[-1][1] - pt[1]) > eps:
                    cleaned.append(pt)

        # Remove closing duplication if any
        if len(cleaned) >= 2 and same_pt(cleaned[0], cleaned[-1]):
            cleaned.pop()

        if len(cleaned) < 3:
            return None

        cleaned = simplify_polygon_collinear(cleaned, eps=params.simplify_collinear_eps)
        if len(cleaned) < 3:
            return None

        cleaned = PolygonUtils.ensure_ccw(cleaned)

        if not PolygonUtils.is_simple_polygon(cleaned, timeout_seconds=2.0):
            return None
        if not PolygonUtils.is_convex(cleaned):
            return None

        return cleaned

    def polygon_ok_inside(convex_poly: Polygon, original_poly: Polygon) -> bool:
        for v in convex_poly:
            if not PolygonUtils.is_point_inside_polygon(v, original_poly):
                return False
        for i in range(len(convex_poly)):
            a = convex_poly[i]
            b = convex_poly[(i + 1) % len(convex_poly)]
            for s in (0.25, 0.5, 0.75):
                p = (a[0] * (1 - s) + b[0] * s, a[1] * (1 - s) + b[1] * s)
                if not PolygonUtils.is_point_inside_polygon(p, original_poly):
                    return False
        return True

    def polygon_area_abs(p: Polygon) -> float:
        return abs(GeometryUtils.calculate_signed_area(p)) * 0.5

    def polygon_min_internal_angle(poly: Polygon) -> float:
        m = len(poly)
        amin = math.pi
        for i in range(m):
            p0 = poly[(i - 1) % m]
            p1 = poly[i]
            p2 = poly[(i + 1) % m]
            v1 = (p0[0] - p1[0], p0[1] - p1[1])
            v2 = (p2[0] - p1[0], p2[1] - p1[1])
            d1 = math.hypot(*v1)
            d2 = math.hypot(*v2)
            if d1 * d2 < 1e-12:
                continue
            c = (v1[0] * v2[0] + v1[1] * v2[1]) / (d1 * d2)
            c = max(-1.0, min(1.0, c))
            amin = min(amin, math.acos(c))
        return amin

    def score_merge(polys: List[Polygon], i: int, j: int, merged: Polygon) -> float:
        ars = [polygon_aspect_ratio_pca(p) for p in polys]
        areas = [polygon_area_abs(p) for p in polys]
        ar_m = polygon_aspect_ratio_pca(merged)
        A_m = polygon_area_abs(merged)

        new_ars = [ar_m] + [ars[k] for k in range(len(polys)) if k not in (i, j)]
        new_areas = [A_m] + [areas[k] for k in range(len(polys)) if k not in (i, j)]

        mm = sum(new_areas) / len(new_areas)
        cv = 0.0 if mm <= 1e-30 else math.sqrt(sum((v - mm) ** 2 for v in new_areas) / len(new_areas)) / mm
        mn, mx = min(new_areas), max(new_areas)
        ratio = float("inf") if mn <= 1e-30 else mx / mn

        worst_ar = max(new_ars)
        min_ang = polygon_min_internal_angle(merged)
        angle_pen = max(0.0, math.radians(30.0) - min_ang)

        if worst_ar <= params.ar_limit:
            return (
                worst_ar
                + params.area_balance_weight * cv
                + params.area_ratio_weight * ratio
                + 6.0 * angle_pen
            )

        if not params.soft_fallback:
            return float("inf")

        viol = worst_ar / params.ar_limit
        return (
            1000.0 * max(0.0, viol - 1.0)
            + 50.0 * cv
            + 10.0 * ratio
            + 10.0 * angle_pen
        )

    while True:
        edge_map = build_edge_map(parts)
        best = None

        for edge, ids in edge_map.items():
            if len(ids) < 2:
                continue
            ids2 = list(dict.fromkeys(ids))
            if len(ids2) != 2:
                continue
            i, j = ids2[0], ids2[1]

            merged = merge_two_convex(parts[i], parts[j])
            if merged is None:
                continue
            if not polygon_ok_inside(merged, poly0):
                continue

            score = score_merge(parts, i, j, merged)
            if not math.isfinite(score):
                continue
            if best is None or score < best[0]:
                best = (score, i, j, merged)

        if best is None:
            break

        _, i, j, merged = best
        new_parts: List[Polygon] = []
        for k, p in enumerate(parts):
            if k in (i, j):
                continue
            new_parts.append(p)
        new_parts.append(merged)
        parts = new_parts

    parts = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=params.simplify_collinear_eps)) for p in parts]
    parts = [p for p in parts if len(p) >= 3 and PolygonUtils.is_convex(p)]
    parts.sort(key=lambda p: -abs(GeometryUtils.calculate_signed_area(p)))
    return parts


def convex_seed_by_y_symmetric_chords(polygon: Polygon, params: PartitionParams) -> List[Polygon]:
    poly = PolygonUtils.ensure_ccw(polygon)
    epsy = float(params.ysym_eps)
    eps = 10 ** (-PRECISION)

    n = len(poly)
    i_left = min(range(n), key=lambda i: (poly[i][0], abs(poly[i][1])))
    i_right = max(range(n), key=lambda i: (poly[i][0], -abs(poly[i][1])))

    def walk(i0: int, i1: int) -> List[int]:
        out = [i0]
        i = i0
        while i != i1:
            i = (i + 1) % n
            out.append(i)
            if len(out) > n + 5:
                break
        return out

    idx_lr = walk(i_left, i_right)
    idx_rl = walk(i_right, i_left)

    def chain_score(idxs: List[int]) -> float:
        return sum(1.0 for i in idxs if poly[i][1] >= -epsy)

    upper_idxs = idx_lr if chain_score(idx_lr) >= chain_score(idx_rl) else idx_rl
    lower_idxs = idx_rl if upper_idxs is idx_lr else idx_lr

    upper_all = [poly[i] for i in upper_idxs]
    lower_all = [poly[i] for i in lower_idxs]

    upper = [p for p in upper_all if p[1] >= -epsy]
    lower = [p for p in lower_all if p[1] <= epsy]

    if not upper or not lower:
        return TriangulationUtils.ear_clipping_triangulation(poly)

    if upper[0][0] > upper[-1][0]:
        upper = list(reversed(upper))
    if lower[0][0] > lower[-1][0]:
        lower = list(reversed(lower))

    def keyx(p: Point) -> float:
        return round(p[0], PRECISION)

    up_by_x: Dict[float, Point] = {}
    lo_by_x: Dict[float, Point] = {}

    for p in upper:
        k = keyx(p)
        if k not in up_by_x or p[1] > up_by_x[k][1]:
            up_by_x[k] = p

    for p in lower:
        k = keyx(p)
        if k not in lo_by_x or p[1] < lo_by_x[k][1]:
            lo_by_x[k] = p

    xs = sorted(set(up_by_x.keys()) & set(lo_by_x.keys()))
    if len(xs) < 2:
        return TriangulationUtils.ear_clipping_triangulation(poly)

    up_list = [up_by_x[xk] for xk in xs]
    lo_list = [lo_by_x[xk] for xk in xs]

    parts: List[Polygon] = []

    for k in range(len(xs) - 1):
        u0 = up_list[k]
        u1 = up_list[k + 1]
        l1 = lo_list[k + 1]
        l0 = lo_list[k]

        if GeometryUtils.distance(u0, l0) <= eps:
            tri = [u0, u1, l1]
            tri = simplify_polygon_collinear(tri, eps=1e-12)
            if len(tri) >= 3:
                tri = PolygonUtils.ensure_ccw(tri)
                if PolygonUtils.is_convex(tri):
                    parts.append(tri)
                else:
                    parts.extend(TriangulationUtils.ear_clipping_triangulation(tri))
            continue

        quad = [u0, u1, l1, l0]
        quad = simplify_polygon_collinear(quad, eps=params.simplify_collinear_eps)
        if len(quad) >= 3:
            quad = PolygonUtils.ensure_ccw(quad)
            if PolygonUtils.is_convex(quad):
                parts.append(quad)
            else:
                parts.extend(TriangulationUtils.ear_clipping_triangulation(quad))

    left_caps = build_cap_polygons(poly, up_list[0], lo_list[0], up_list[1], lo_list[1], side="left", eps=eps)
    right_caps = build_cap_polygons(poly, up_list[-1], lo_list[-1], up_list[-2], lo_list[-2], side="right", eps=eps)

    if left_caps:
        parts.extend(left_caps)
    if right_caps:
        parts.extend(right_caps)

    parts = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=params.simplify_collinear_eps)) for p in parts]
    parts = [p for p in parts if len(p) >= 3 and PolygonUtils.is_simple_polygon(p, timeout_seconds=1.5)]
    parts = [p for p in parts if PolygonUtils.is_convex(p)]

    if not parts:
        return TriangulationUtils.ear_clipping_triangulation(poly)

    return parts


def load_polygon_vertices_from_excel(xlsx_path: str, sheet_name: str = "Outer_Polygon") -> Polygon:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    vertices: Polygon = []
    for row in rows[1:]:
        if row is None or len(row) < 2 or row[0] is None or row[1] is None:
            break
        vertices.append((float(row[0]), float(row[1])))

    if len(vertices) < 3:
        raise ValueError("Polygon vertex count < 3 in Excel.")
    return vertices


def plot_partition(original: Polygon, parts: List[Polygon]) -> None:
    plt.figure(figsize=(10, 8))

    ox, oy = zip(*original)
    plt.plot(ox + (ox[0],), oy + (oy[0],), "k--", lw=2, label="Outer Polygon")

    colors = plt.cm.tab20.colors

    for i, p in enumerate(parts):
        px, py = zip(*p)
        plt.fill(px, py, color=colors[i % len(colors)], alpha=0.25)
        plt.plot(px + (px[0],), py + (py[0],), "-", lw=2, color=colors[i % len(colors)])

    plt.title(f"Convex Partition (parts={len(parts)})")
    plt.axis("equal")
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.show()


def polygon_aspect_ratio_pca(polygon: Polygon, eps: float = 1e-12) -> float:
    import numpy as np

    P = np.asarray(polygon, dtype=float)
    if P.shape[0] < 3:
        return float("inf")

    C = P.mean(axis=0, keepdims=True)
    X = P - C
    cov = (X.T @ X) / max(P.shape[0], 1)

    w = np.linalg.eigvalsh(cov)
    w0 = float(max(w[0], eps))
    w1 = float(max(w[-1], eps))
    return math.sqrt(w1 / w0)


def _copy_sheet_with_format(src_wb, dest_wb, sheet_name: str) -> None:
    if sheet_name not in src_wb.sheetnames:
        return
    if sheet_name in dest_wb.sheetnames:
        del dest_wb[sheet_name]
    src_ws = src_wb[sheet_name]
    dest_ws = dest_wb.create_sheet(title=sheet_name)
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))


def save_partition_to_excel(
    output_xlsx: str,
    original: Polygon,
    parts: List[Polygon],
    params: PartitionParams,
    source_xlsx: str,
) -> None:
    wb = Workbook()
    ws = wb.active

    # 主表名：尽量沿用你现在的参数命名；没有就给一个默认值
    ws.title = "outer_polygon_lobular2"

    # 1) 写入 Original Polygon（与你的格式一致）
    ws.append(["Original Polygon X", "Original Polygon Y"])
    for x, y in original:
        ws.append([float(x), float(y)])

    # 2) 写入 Convex Parts（空行分段 + 指定表头）
    for i, part in enumerate(parts, start=1):
        ws.append([])
        ws.append([f"Convex Part {i} X", f"Convex Part {i} Y"])
        for x, y in part:
            ws.append([float(x), float(y)])

    # 3) Metadata sheet（与你的格式一致：Attribute / Value）
    meta = wb.create_sheet(title="Metadata")
    meta.append(["Attribute", "Value"])

    # 统计：凹点数量（对 original 做一个稳健计算）
    def _count_concave_vertices(poly: Polygon, eps: float = 1e-12) -> int:
        poly_ccw = PolygonUtils.ensure_ccw(poly)
        n = len(poly_ccw)
        if n < 3:
            return 0
        cnt = 0
        for k in range(n):
            a = poly_ccw[(k - 1) % n]
            b = poly_ccw[k]
            c = poly_ccw[(k + 1) % n]
            if GeometryUtils.cross(a, b, c) < -eps:
                cnt += 1
        return cnt

    # 中心（与你示例保持一致：简单平均）
    cx = sum(v[0] for v in original) / max(len(original), 1)
    cy = sum(v[1] for v in original) / max(len(original), 1)

    meta.append(["Concave Points", _count_concave_vertices(original)])
    meta.append(["Convex Parts", len(parts)])
    meta.append(["Center X", float(cx)])
    meta.append(["Center Y", float(cy)])
    meta.append(["Original Vertices", len(original)])
    meta.append(["ArLimit", float(getattr(params, "ar_limit", 0.0))])
    meta.append(["SoftFallback", int(bool(getattr(params, "soft_fallback", False)))])

    source_wb = load_workbook(source_xlsx)
    _copy_sheet_with_format(source_wb, wb, "curve_pea")
    _copy_sheet_with_format(source_wb, wb, "area_pea")
    _copy_sheet_with_format(source_wb, wb, "curve_lobular2")
    _copy_sheet_with_format(source_wb, wb, "area_lobular2")

    wb.save(output_xlsx)


def simplify_polygon_collinear(polygon: Polygon, eps: float = 1e-10) -> Polygon:
    poly = list(polygon)
    if len(poly) < 3:
        return poly

    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        keep: List[Point] = []
        for i in range(n):
            a = poly[(i - 1) % n]
            b = poly[i]
            c = poly[(i + 1) % n]
            if GeometryUtils.distance(a, b) < eps:
                changed = True
                continue
            if abs(GeometryUtils.cross(a, b, c)) <= eps and (
                min(a[0], c[0]) - eps <= b[0] <= max(a[0], c[0]) + eps and
                min(a[1], c[1]) - eps <= b[1] <= max(a[1], c[1]) + eps
            ):
                changed = True
                continue
            keep.append(b)
        poly = keep

    if len(poly) < 3:
        return list(polygon)
    return poly


def _normalize_circumscribe_mode(mode: str) -> str:
    m = str(mode).strip().lower()
    if m in ("general", "gen", "generic"):
        return "general"
    if m in ("symmetric", "sym", "duicheng", "对称"):
        return "symmetric"
    raise ValueError("circumscribe_mode must be 'symmetric' or 'general'.")


def concave_decomp_lobular2(
    scale_ratio_collision: float = 1.025,
    num_vertices: int = 16,
    plot: bool = True,
    circumscribe_mode: str = "symmetric",
    general_num_points: int = 240,
) -> None:
    plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False

    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    input_xlsx = data_dir / "outer_polygon_vertices_lobular2.xlsx"
    output_xlsx = data_dir / "convex_partition_lobular2.xlsx"
    mode = _normalize_circumscribe_mode(circumscribe_mode)
    if mode == "general":
        from shape_decomposition.circum_concave_general import circum_concave_general
        circum_concave_general(
            output_xlsx=input_xlsx,
            scale_ratio_collision=scale_ratio_collision,
            shape_name="lobular2",
            num_vertices=num_vertices,
            num_points=general_num_points,
            plot=plot,
        )
    else:
        from shape_decomposition.circum_lobular2 import circum_lobular2
        circum_lobular2(
            output_xlsx=input_xlsx,
            scale_ratio_collision=scale_ratio_collision,
            num_vertices=num_vertices,
            plot=plot,
        )

    params = PartitionParams(
        ar_limit=3.0,
        area_balance_weight=2.5,
        area_ratio_weight=0.8,
        plot=bool(plot),
        sheet_name="outer_polygon_lobular2",
        simplify_collinear_eps=1e-10,
        soft_fallback=True,
        ysym_eps=1e-6,
    )

    poly = load_polygon_vertices_from_excel(input_xlsx, sheet_name=params.sheet_name)
    poly = PolygonUtils.ensure_ccw(poly)

    if not PolygonUtils.is_simple_polygon(poly):
        raise ValueError("Input polygon is not simple (self-intersecting). Please fix Excel vertices order/data.")

    parts = convex_partition(poly, params)

    if params.plot:
        plot_partition(poly, parts)

    save_partition_to_excel(
        output_xlsx=output_xlsx,
        original=poly,
        parts=parts,
        params=params,
        source_xlsx=input_xlsx,
    )


def main() -> None:
    concave_decomp_lobular2()


if __name__ == "__main__":
    main()
