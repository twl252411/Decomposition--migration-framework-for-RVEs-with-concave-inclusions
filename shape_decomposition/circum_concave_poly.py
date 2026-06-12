# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from typing import List, Tuple
from openpyxl import Workbook
import numpy as np

Point = Tuple[float, float]


DEFAULT_VERTICES: List[Point] = [
    (-8.7611369999999997, -0.74006099999999997),
    (-7.5917560000000002, 1.566219),
    (-3.2715420000000002, 4.1323610000000004),
    (1.9581900000000001, 4.0673950000000003),
    (3.1925370000000002, 5.2367759999999999),
    (7.9999929999999999, 4.1323610000000004),
    (8.5522010000000002, 0.98152799999999996),
    (2.8677090000000001, -2.6240640000000002),
    (-2.6868509999999999, -8.3085559999999994),
    (-3.3689909999999998, -7.6589),
    (-4.0186469999999996, -1.5521309999999999),
    (-6.7472029999999998, -2.006891),
]


def _translate_to_center(points: List[Point]) -> List[Point]:
    if not points:
        return []
    center_x = sum(x for x, _ in points) / len(points)
    center_y = sum(y for _, y in points) / len(points)
    return [(x - center_x, y - center_y) for x, y in points]


def _polygon_area(points: List[Point]) -> float:
    if len(points) < 3:
        return 0.0
    area = 0.0
    n = len(points)
    for i in range(n):
        j = (i + 1) % n
        area += points[i][0] * points[j][1] - points[j][0] * points[i][1]
    return abs(area) * 0.5


def save_outer_polygon_to_excel(
    vertices_scaled: List[Point],
    vertices_unscaled: List[Point],
    area_unscaled: float,
    output_xlsx: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "outer_polygon_concave_poly"
    ws.append(["X", "Y"])
    for x, y in vertices_scaled:
        ws.append([float(x), float(y)])

    ws2 = wb.create_sheet(title="curve_concave_poly")
    ws2.append(["X", "Y"])
    for x, y in vertices_unscaled:
        ws2.append([float(x), float(y)])

    ws3 = wb.create_sheet(title="area_concave_poly")
    ws3.append(["area_curve"])
    ws3.append([float(area_unscaled)])
    wb.save(output_xlsx)


def circum_concave_poly(output_xlsx, scale_ratio_collision: float = 1.025) -> List[Point]:
    translated = _translate_to_center(DEFAULT_VERTICES)
    scaled = [(x * scale_ratio_collision, y * scale_ratio_collision) for x, y in translated]
    area_unscaled = _polygon_area(translated)
    save_outer_polygon_to_excel(scaled, translated, area_unscaled, str(output_xlsx))

    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    output_txt = data_dir / "vertices_concave_poly.txt"
    np.savetxt(output_txt, np.array(translated, dtype=float), fmt="%.10f")
    return scaled


def main() -> None:
    scale_ratio_collision = 1.025
    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = data_dir / "outer_polygon_vertices_concave_poly.xlsx"
    circum_concave_poly(output_xlsx=output_xlsx, scale_ratio_collision=scale_ratio_collision)


if __name__ == "__main__":
    main()
