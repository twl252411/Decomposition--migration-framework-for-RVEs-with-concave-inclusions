from pathlib import Path
from math import pi

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.path import Path as MplPath
from openpyxl import Workbook


def close_polyline(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if not (np.isclose(x[0], x[-1]) and np.isclose(y[0], y[-1])):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]
    return x, y


def concave_outer_polygon_xsym(
    P_curve_dense,
    num_vertices=16,
    offset_ratio=0.003,
    bisect_iters=60,
    refine=True,
    refine_passes=3,
    refine_bisect=28,
    endpoint_bisect=32,
    enforce_area_monotonic=True,
):
    """
    x 轴对称的外接多边形（由上链 + 镜像下链构造），并保证外接原曲线点集。
    关键特性：
      - 先统一 eps 外偏置找到最小可行
      - 再逐点 refine eps（除端点）
      - 最后端点沿 x 轴进一步贴近（alpha 二分），但同时保证“无自交 + 外接”
    """
    if num_vertices < 8:
        raise ValueError("Use >= 8 for stable concave approximation.")
    if num_vertices % 2 != 0:
        raise ValueError("For x-symmetry, num_vertices must be even.")

    P = np.asarray(P_curve_dense, float)
    if not np.allclose(P[0], P[-1]):
        P = np.vstack([P, P[0]])

    scale = float(np.hypot(np.ptp(P[:, 0]), np.ptp(P[:, 1])) + 1e-12)
    centroid = np.mean(P[:-1], axis=0)

    # -------------------------
    # local helpers (self-intersection)
    # -------------------------
    def _seg_intersect(a, b, c, d, eps=1e-12):
        def orient(p, q, r):
            return (q[0] - p[0]) * (r[1] - p[1]) - (q[1] - p[1]) * (r[0] - p[0])

        def onseg(p, q, r):
            return (
                min(p[0], r[0]) - eps <= q[0] <= max(p[0], r[0]) + eps
                and min(p[1], r[1]) - eps <= q[1] <= max(p[1], r[1]) + eps
            )

        o1 = orient(a, b, c)
        o2 = orient(a, b, d)
        o3 = orient(c, d, a)
        o4 = orient(c, d, b)

        if (o1 * o2 < -eps) and (o3 * o4 < -eps):
            return True
        if abs(o1) <= eps and onseg(a, c, b):
            return True
        if abs(o2) <= eps and onseg(a, d, b):
            return True
        if abs(o3) <= eps and onseg(c, a, d):
            return True
        if abs(o4) <= eps and onseg(c, b, d):
            return True
        return False

    def _is_simple_polygon(V):
        V = np.asarray(V, float)
        n = V.shape[0]
        for i in range(n):
            a = V[i]
            b = V[(i + 1) % n]
            for j in range(i + 1, n):
                # 跳过共享端点的相邻边
                if j == i:
                    continue
                if (j == i + 1) or ((i == 0) and (j == n - 1)):
                    continue
                if (i == j + 1) or ((j == 0) and (i == n - 1)):
                    continue
                c = V[j]
                d = V[(j + 1) % n]
                if _seg_intersect(a, b, c, d):
                    return False
        return True

    def _poly_intersects_curve(V, curve, eps=1e-12):
        V = np.asarray(V, float)
        C = np.asarray(curve, float)
        if C.shape[0] < 2:
            return False
        if not np.allclose(C[0], C[-1]):
            C = np.vstack([C, C[0]])
        n = V.shape[0]
        m = C.shape[0] - 1
        for i in range(n):
            a = V[i]
            b = V[(i + 1) % n]
            for j in range(m):
                c = C[j]
                d = C[j + 1]
                if _seg_intersect(a, b, c, d, eps=eps):
                    return True
        return False

    # -------------------------
    # 1) 提取连续上半段 + 端点补 y=0
    # -------------------------
    y = P[:-1, 1]
    m = y.size
    up = y >= 0.0
    up2 = np.r_[up, up]

    best_len = 0
    best_end = -1
    cur = 0
    for i in range(up2.size):
        if up2[i]:
            cur += 1
            if cur > best_len:
                best_len = cur
                best_end = i
        else:
            cur = 0
    if best_len < 2:
        raise RuntimeError("Upper chain extraction failed.")

    best_start = best_end - best_len + 1
    idx = (np.arange(best_start, best_end + 1) % m).astype(int)
    P_up_raw = P[idx]

    def interp_y0(A, B):
        y1, y2 = A[1], B[1]
        if abs(y2 - y1) < 1e-15:
            return A.copy()
        t = (0.0 - y1) / (y2 - y1)
        return A + t * (B - A)

    i0 = idx[0]
    iprev = (i0 - 1) % m
    A = P[iprev]
    B = P[i0]
    P0 = interp_y0(A, B) if (A[1] < 0.0 and B[1] >= 0.0) else B.copy()
    P0[1] = 0.0

    i1 = idx[-1]
    inext = (i1 + 1) % m
    A = P[i1]
    B = P[inext]
    P1 = interp_y0(A, B) if (A[1] >= 0.0 and B[1] < 0.0) else A.copy()
    P1[1] = 0.0

    P_up = np.vstack([P0, P_up_raw, P1])
    if P_up[0, 0] > P_up[-1, 0]:
        P_up = P_up[::-1].copy()

    # 在可能翻转后再记录端点基准
    x0_base = float(P_up[0, 0])
    x1_base = float(P_up[-1, 0])

    # -------------------------
    # 2) 上半段重采样 + 外法线
    # -------------------------
    m_up = num_vertices // 2 + 1
    P_up_rs = resample_open_polyline_by_arclength(P_up, m_up)
    N_up = outward_normals(P_up_rs, centroid)

    # -------------------------
    # 3) 统一 eps 找最小可行外偏置
    # -------------------------
    lo = 0.0
    hi = float(max(offset_ratio, 1e-6) * scale)

    def build_poly_from_epsvec(eps_vec_up, alpha0=1.0, alpha1=1.0):
        Q_up = P_up_rs + eps_vec_up[:, None] * N_up

        Q_up[0, 1] = 0.0
        Q_up[-1, 1] = 0.0

        x0_out = float(Q_up[0, 0])
        x1_out = float(Q_up[-1, 0])
        Q_up[0, 0] = x0_base + float(alpha0) * (x0_out - x0_base)
        Q_up[-1, 0] = x1_base + float(alpha1) * (x1_out - x1_base)

        Q_low = Q_up[-2:0:-1].copy()
        Q_low[:, 1] *= -1.0

        V = np.vstack([Q_up, Q_low])
        if poly_area_signed(V) < 0:
            V = V[::-1].copy()
        return V

    def ok_poly(V):
        if not _is_simple_polygon(V):
            return False
        path = MplPath(V, closed=True)
        inside = path.contains_points(P[:-1], radius=1e-10 * scale)
        if not bool(np.all(inside)):
            return False
        if _poly_intersects_curve(V, P, eps=1e-12):
            return False
        return True

    def ok_uniform(eps):
        eps_vec = np.full(m_up, eps, float)
        V = build_poly_from_epsvec(eps_vec, alpha0=1.0, alpha1=1.0)
        return ok_poly(V)

    grow = 0
    while not ok_uniform(hi) and grow < 80:
        lo = hi
        hi *= 1.6
        grow += 1
    if not ok_uniform(hi):
        eps_vec = np.full(m_up, hi, float)
        poly_final = build_poly_from_epsvec(eps_vec, alpha0=1.0, alpha1=1.0)
        if enforce_area_monotonic and num_vertices >= 10:
            poly_prev = concave_outer_polygon_xsym(
                P_curve_dense=P,
                num_vertices=num_vertices - 2,
                offset_ratio=offset_ratio,
                bisect_iters=bisect_iters,
                refine=refine,
                refine_passes=refine_passes,
                refine_bisect=refine_bisect,
                endpoint_bisect=endpoint_bisect,
                enforce_area_monotonic=True,
            )
            poly_prev_expanded = expand_xsym_polygon_vertices(poly_prev)
            area_prev = polygon_area(poly_prev)
            if polygon_area(poly_final) > area_prev + 1e-10 * scale * scale:
                poly_final = poly_prev_expanded
        return poly_final

    for _ in range(int(bisect_iters)):
        mid = 0.5 * (lo + hi)
        if ok_uniform(mid):
            hi = mid
        else:
            lo = mid

    eps_vec = np.full(m_up, hi, float)

    # -------------------------
    # 4) 逐点收缩 eps_vec（端点固定）
    # -------------------------
    if refine:
        fixed = np.zeros(m_up, dtype=bool)
        fixed[0] = True
        fixed[-1] = True

        for _pass in range(int(refine_passes)):
            improved = False
            for i in range(m_up):
                if fixed[i]:
                    continue

                lo_i = 0.0
                hi_i = float(eps_vec[i])

                trial = eps_vec.copy()
                trial[i] = 0.0
                if ok_poly(build_poly_from_epsvec(trial, alpha0=1.0, alpha1=1.0)):
                    eps_vec[i] = 0.0
                    improved = True
                    continue

                for _ in range(int(refine_bisect)):
                    mid = 0.5 * (lo_i + hi_i)
                    trial = eps_vec.copy()
                    trial[i] = mid
                    if ok_poly(build_poly_from_epsvec(trial, alpha0=1.0, alpha1=1.0)):
                        hi_i = mid
                    else:
                        lo_i = mid

                if hi_i < eps_vec[i] - 1e-12 * scale:
                    eps_vec[i] = hi_i
                    improved = True

            if not improved:
                break

    # -------------------------
    # 5) 端点 alpha 二分贴近（只动 x，y=0 固定），并保证无自交+外接
    # -------------------------
    alpha0 = 1.0
    alpha1 = 1.0

    def ok_alpha(a0, a1):
        return ok_poly(build_poly_from_epsvec(eps_vec, alpha0=a0, alpha1=a1))

    lo_a, hi_a = 0.0, 1.0
    if ok_alpha(lo_a, alpha1):
        alpha0 = lo_a
    else:
        for _ in range(int(endpoint_bisect)):
            mid = 0.5 * (lo_a + hi_a)
            if ok_alpha(mid, alpha1):
                hi_a = mid
            else:
                lo_a = mid
        alpha0 = hi_a

    lo_a, hi_a = 0.0, 1.0
    if ok_alpha(alpha0, lo_a):
        alpha1 = lo_a
    else:
        for _ in range(int(endpoint_bisect)):
            mid = 0.5 * (lo_a + hi_a)
            if ok_alpha(alpha0, mid):
                hi_a = mid
            else:
                lo_a = mid
        alpha1 = hi_a

    poly_final = build_poly_from_epsvec(eps_vec, alpha0=alpha0, alpha1=alpha1)

    if enforce_area_monotonic and num_vertices >= 10:
        poly_prev = concave_outer_polygon_xsym(
            P_curve_dense=P,
            num_vertices=num_vertices - 2,
            offset_ratio=offset_ratio,
            bisect_iters=bisect_iters,
            refine=refine,
            refine_passes=refine_passes,
            refine_bisect=refine_bisect,
            endpoint_bisect=endpoint_bisect,
            enforce_area_monotonic=True,
        )
        poly_prev_expanded = expand_xsym_polygon_vertices(poly_prev)
        area_prev = polygon_area(poly_prev)

        if polygon_area(poly_final) > area_prev + 1e-10 * scale * scale:
            poly_final = poly_prev_expanded

    return poly_final


def expand_xsym_polygon_vertices(poly):
    """
    Add one midpoint to the upper chain and mirror it to the lower chain,
    increasing the vertex count by 2 while preserving the polygon boundary.
    """
    V = np.asarray(poly, float)
    n = V.shape[0]
    if n < 8 or n % 2 != 0:
        raise ValueError("poly must have an even number of vertices and n >= 8.")

    up = _extract_upper_chain_xsym(V)

    seg_len = np.hypot(np.diff(up[:, 0]), np.diff(up[:, 1]))
    i = int(np.argmax(seg_len))
    mid = 0.5 * (up[i] + up[i + 1])

    up_new = np.insert(up, i + 1, mid, axis=0)
    return _build_xsym_polygon_from_upper(up_new)


def _build_xsym_polygon_from_upper(up):
    up = np.asarray(up, float)
    low = up[-2:0:-1].copy()
    low[:, 1] *= -1.0
    V = np.vstack([up, low])
    if poly_area_signed(V) < 0:
        V = V[::-1].copy()
    return V


def _extract_upper_chain_xsym(poly):
    V = np.asarray(poly, float)
    n = V.shape[0]
    tol = 1e-10 * (np.ptp(V[:, 0]) + np.ptp(V[:, 1]) + 1.0)

    axis_idx = np.where(np.abs(V[:, 1]) <= tol)[0]
    if axis_idx.size < 2:
        raise ValueError("Cannot identify x-axis endpoints for x-symmetric polygon.")

    i_left = int(axis_idx[np.argmin(V[axis_idx, 0])])
    i_right = int(axis_idx[np.argmax(V[axis_idx, 0])])
    if i_left == i_right:
        raise ValueError("Degenerate x-axis endpoints.")

    def walk(a, b, step):
        idx = [a]
        i = a
        while i != b:
            i = (i + step) % n
            idx.append(i)
        return np.asarray(idx, dtype=int)

    path_fwd = walk(i_left, i_right, +1)
    path_bwd = walk(i_left, i_right, -1)

    y_fwd = np.mean(V[path_fwd, 1])
    y_bwd = np.mean(V[path_bwd, 1])
    up_idx = path_fwd if y_fwd >= y_bwd else path_bwd

    up = V[up_idx].copy()
    if up[0, 0] > up[-1, 0]:
        up = up[::-1].copy()
    up[0, 1] = 0.0
    up[-1, 1] = 0.0
    return up


def create_kidney_shape(R1=12.0, R2=10.0, K1=15.0, a=1.5, num_points=240):
    """
    x = R1 cos t + K1 exp(-a cos t - a)
    y = R2 sin t
    """
    t = np.linspace(0.0, 2 * pi, num_points, endpoint=False)
    x = R1 * np.cos(t) + K1 * np.exp(-a * np.cos(t) - a)
    y = R2 * np.sin(t)
    return x, y, t


def curve_area_shoelace(x, y):
    x, y = close_polyline(x, y)
    return 0.5 * abs(np.sum(x[:-1] * y[1:] - x[1:] * y[:-1]))


def outward_normals(P, centroid):
    """
    Compute per-vertex outward normals for an OPEN polyline P (ordered).
    Normal direction chosen by dot(n, p-centroid) > 0.
    """
    P = np.asarray(P, float)
    T = np.zeros_like(P)
    T[1:-1] = P[2:] - P[:-2]
    T[0] = P[1] - P[0]
    T[-1] = P[-1] - P[-2]

    tn = np.hypot(T[:, 0], T[:, 1]) + 1e-12
    T = T / tn[:, None]

    N = np.column_stack([-T[:, 1], T[:, 0]])
    v = P - centroid[None, :]
    sgn = np.sign(np.sum(N * v, axis=1))
    sgn[sgn == 0] = 1.0
    return N * sgn[:, None]


def polygon_area(v):
    v = np.asarray(v, float)
    x = v[:, 0]
    y = v[:, 1]
    return 0.5 * abs(np.sum(x * np.roll(y, -1) - np.roll(x, -1) * y))


def poly_area_signed(V):
    x = V[:, 0]
    y = V[:, 1]
    return 0.5 * np.sum(x * np.roll(y, -1) - np.roll(x, -1) * y)


def resample_open_polyline_by_arclength(P, m):
    P = np.asarray(P, float)
    if P.shape[0] < 2:
        return P.copy()
    d = np.hypot(np.diff(P[:, 0]), np.diff(P[:, 1])) + 1e-12
    s = np.r_[0.0, np.cumsum(d)]
    L = float(s[-1])
    t = np.linspace(0.0, L, m)
    x = np.interp(t, s, P[:, 0])
    y = np.interp(t, s, P[:, 1])
    return np.column_stack([x, y])


def resample_polyline_by_arclength(P, m):
    """
    Resample an (optionally closed) polyline P to m points by arclength.
    If P is closed (first==last), it keeps closure.
    """
    P = np.asarray(P, float)
    if P.shape[0] < 2:
        return P.copy()

    closed = np.allclose(P[0], P[-1], atol=1e-12)
    if not closed:
        P = np.vstack([P, P[0]])
        closed = True

    d = np.hypot(np.diff(P[:, 0]), np.diff(P[:, 1])) + 1e-12
    s = np.r_[0.0, np.cumsum(d)]
    L = float(s[-1])

    if L <= 1e-12:
        idx = np.linspace(0, P.shape[0] - 1, m).round().astype(int)
        Q = P[idx]
        if closed:
            Q[-1] = Q[0]
        return Q

    if closed:
        t = np.linspace(0.0, L, m, endpoint=False)
    else:
        t = np.linspace(0.0, L, m)

    x = np.interp(t, s, P[:, 0])
    y = np.interp(t, s, P[:, 1])
    Q = np.column_stack([x, y])

    if closed:
        Q = np.vstack([Q, Q[0]])
    return Q


def save_outer_polygon_to_excel(poly, a_poly, x, y, a_curve,
    filename="outer_polygon_vertices.xlsx", include_closed=False):
    """
    Save outer polygon vertices to Excel.

    Parameters
    ----------
    poly : array-like of shape (N, 2)
        Polygon vertices (not necessarily closed).
    filename : str
        Output Excel filename.
    sheet_title : str
        Worksheet name.
    include_closed : bool
        If True, append the first vertex at the end to close the polygon in Excel.
    """
    V = np.asarray(poly, dtype=float)
    if V.ndim != 2 or V.shape[1] != 2 or V.shape[0] < 3:
        raise ValueError("poly must be an array-like of shape (N, 2) with N >= 3.")

    if include_closed and (not np.allclose(V[0], V[-1], atol=1e-12)):
        V = np.vstack([V, V[0]])

    wb = Workbook()
    ws = wb.active
    sheet_title = "outer_polygon_pea"
    ws.title = sheet_title
    ws.append(["X", "Y"])
    for xi, yi in V:
        ws.append([float(xi), float(yi)])

    ws2 = wb.create_sheet(title="curve_pea")
    ws2.append(["X", "Y"])
    for xi, yi in zip(x, y):
        ws2.append([float(xi), float(yi)])

    ws3 = wb.create_sheet(title="area_pea")
    ws3.append(["area_poly", "area_curve"])
    ws3.append([float(a_poly), float(a_curve)])

    wb.save(filename)


def x_axis_intersection_center(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if x.size == 0:
        raise ValueError("Curve vertices are empty.")
    if x.size < 2:
        return float(x[0]), 0.0
    if not (np.isclose(x[0], x[-1]) and np.isclose(y[0], y[-1])):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]

    xs = []
    for i in range(x.size - 1):
        x1, y1 = float(x[i]), float(y[i])
        x2, y2 = float(x[i + 1]), float(y[i + 1])
        if np.isclose(y1, 0.0):
            xs.append(x1)
        if np.isclose(y2, 0.0):
            xs.append(x2)
        if (y1 > 0.0 and y2 < 0.0) or (y1 < 0.0 and y2 > 0.0):
            t = -y1 / (y2 - y1)
            xs.append(x1 + t * (x2 - x1))

    if not xs:
        return float(np.mean(x[:-1])), float(np.mean(y[:-1]))

    x_min = min(xs)
    x_max = max(xs)
    return 0.5 * (x_min + x_max), 0.0


def visualize_pea(R1=9.84, R2=8.2, K1=12.3, a=1.4, scale_ratio=1.0,
                       num_vertices=16, num_points=240, output_xlsx="", plot: bool = True):
    plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False

    # 1) kidney curve
    x, y, t = create_kidney_shape(R1=R1, R2=R2, K1=K1, a=a, num_points=num_points)
    center_x, center_y = x_axis_intersection_center(x, y)
    x = np.asarray(x, float) - center_x
    y = np.asarray(y, float) - center_y

    # 2) resample by arclength, then scale to target area
    P = np.column_stack([x, y])
    P = np.vstack([P, P[0]])
    P_s = resample_polyline_by_arclength(P, 600)
    xs, ys = P_s[:, 0], P_s[:, 1]

    xs_scaled = xs * scale_ratio
    ys_scaled = ys * scale_ratio
    P_scaled = np.column_stack([xs_scaled, ys_scaled])

    # 3) outer polygon (x-sym)
    poly_scaled = concave_outer_polygon_xsym(
        P_scaled,
        num_vertices=num_vertices,
        offset_ratio=0.002,
        bisect_iters=70,
        refine=True,
        refine_passes=4,
        refine_bisect=30,
        endpoint_bisect=36,
    )
    poly_closed = np.vstack([poly_scaled, poly_scaled[0]])

    # areas
    a_curve = curve_area_shoelace(x, y)
    a_poly = polygon_area(poly_scaled)

    if plot:
        # plot
        plt.figure(figsize=(10, 8))
        plt.plot(xs, ys, "b-", lw=2, label=f"豌豆形（面积：{a_curve:.4f}）")
        plt.plot(
            poly_closed[:, 0],
            poly_closed[:, 1],
            "r--",
            lw=2,
            label=f"外接多边形（面积：{a_poly:.4f}，顶点数：{num_vertices}）",
        )
        plt.plot(poly_scaled[:, 0], poly_scaled[:, 1], "ro", ms=6)

        plt.title(f"豌豆形 及其外接多边形（目标面积 {a_poly}）", fontsize=16)
        plt.xlabel("X")
        plt.ylabel("Y")
        plt.grid(True, ls="--", alpha=0.6)
        plt.gca().set_aspect("equal", adjustable="box")
        plt.legend(loc="upper right")
        plt.show()

    # save
    save_outer_polygon_to_excel(poly_scaled, a_poly, x, y, a_curve,
                                filename=output_xlsx, include_closed=False)
    return poly_scaled, a_poly, x, y, a_curve


def circum_pea(output_xlsx, scale_ratio_collision=1.025, num_vertices=24, plot: bool = True):
    return visualize_pea(
        R1=9.84,
        R2=8.2,
        K1=12.3,
        a=1.4,
        scale_ratio=scale_ratio_collision,
        num_vertices=num_vertices,
        num_points=240,
        output_xlsx=str(output_xlsx),
        plot=plot,
    )

def main():

    scale_ratio_collision = 1.0
    # save
    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = data_dir / "outer_polygon_vertices_pea.xlsx"
    poly = circum_pea(
        output_xlsx=output_xlsx,
        scale_ratio_collision=scale_ratio_collision,
        num_vertices=16,
    )


if __name__ == "__main__":
    main()
