from pathlib import Path
from math import sqrt, pi

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.path import Path as MplPath
from openpyxl import Workbook


# =========================
# shape generators (ONLY: lobular2)
# =========================
def generate_arc(center, start_angle, end_angle, radius, direction="COUNTERCLOCKWISE", num_points=100):
    if direction == "COUNTERCLOCKWISE" and end_angle <= start_angle:
        end_angle += 2 * pi
    elif direction == "CLOCKWISE" and end_angle >= start_angle:
        end_angle -= 2 * pi

    angles = np.linspace(start_angle, end_angle, num_points)
    x = center[0] + radius * np.cos(angles)
    y = center[1] + radius * np.sin(angles)
    return np.column_stack((x, y))


def _rotate_points_90_ccw(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    return -y, x


def _ensure_lobular2_left_right(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if (y.max() - y.min()) > (x.max() - x.min()):
        x, y = _rotate_points_90_ccw(x, y)
    return x, y


def _generate_lobular(arcs, num_points=100):
    all_points = []
    for arc in arcs:
        start_angle = arc["start"]
        end_angle = arc["end"]
        radius = arc["radius"]
        center = arc["center"]

        arc_points = generate_arc(
            center=center,
            start_angle=start_angle,
            end_angle=end_angle,
            radius=radius,
            direction=arc["direction"],
            num_points=arc.get("num_points", num_points),
        )
        all_points.append(arc_points)

    combined = np.concatenate(all_points, axis=0)
    x, y = combined[:, 0], combined[:, 1]
    theta = np.linspace(0, 2 * np.pi, len(x), endpoint=False)
    return x, y, theta


def create_lobular2(radius=4.17284):
    arcs = [
        {"center": (-radius, 0.0), "start": pi / 3, "end": 5 * pi / 3, "radius": radius, "direction": "COUNTERCLOCKWISE"},
        {"center": (0.0, -radius * sqrt(3)), "start": 2 * pi / 3, "end": pi / 3, "radius": radius, "direction": "CLOCKWISE"},
        {"center": (radius, 0.0), "start": 4 * pi / 3, "end": 2 * pi / 3 + 2 * pi, "radius": radius, "direction": "COUNTERCLOCKWISE"},
        {"center": (0.0, radius * sqrt(3)), "start": 5 * pi / 3, "end": 4 * pi / 3, "radius": radius, "direction": "CLOCKWISE", "num_points": 150},
    ]
    x, y, theta = _generate_lobular(arcs)
    x, y = _ensure_lobular2_left_right(x, y)
    return x, y, theta


# =========================
# geometry helpers
# =========================
def _interp_seg(p0, p1, v0, v1, target=0.0):
    den = float(v1 - v0)
    if abs(den) < 1e-15:
        t = 0.0
    else:
        t = float((target - v0) / den)
        t = float(np.clip(t, 0.0, 1.0))
    return p0 + t * (p1 - p0)


def _axis_crossings_positive(x, y, axis="x", eps=1e-14):
    """
    返回正半轴交点 (point, idx_left)
    - axis="x": 求 y=0 且 x>0 的交点（取最大 x）
    - axis="y": 求 x=0 且 y>0 的交点（取最大 y）
    idx_left 是交点所在的线段起点索引 i（线段 i -> i+1）
    """
    x = np.asarray(x, float)
    y = np.asarray(y, float)

    if axis == "x":
        v0 = y[:-1]
        v1 = y[1:]
        idx = np.where(v0 * v1 <= 0.0)[0]
        best = None
        for i in idx:
            if abs(v0[i]) < eps and abs(v1[i]) < eps:
                continue
            xc = _interp_seg(x[i], x[i + 1], v0[i], v1[i], target=0.0)
            if xc > 0.0:
                if (best is None) or (xc > best[0]):
                    best = (float(xc), i)
        if best is None:
            j = int(np.argmax(x))
            return np.array([float(x[j]), 0.0]), j
        return np.array([best[0], 0.0]), int(best[1])

    elif axis == "y":
        v0 = x[:-1]
        v1 = x[1:]
        idx = np.where(v0 * v1 <= 0.0)[0]
        best = None
        for i in idx:
            if abs(v0[i]) < eps and abs(v1[i]) < eps:
                continue
            yc = _interp_seg(y[i], y[i + 1], v0[i], v1[i], target=0.0)
            if yc > 0.0:
                if (best is None) or (yc > best[0]):
                    best = (float(yc), i)
        if best is None:
            j = int(np.argmax(y))
            return np.array([0.0, float(y[j])]), j
        return np.array([0.0, best[0]]), int(best[1])

    else:
        raise ValueError("axis must be 'x' or 'y'")


def _wrap_indices(i0, i1, n, forward=True):
    i0 = int(i0)
    i1 = int(i1)
    n = int(n)
    if forward:
        if i1 >= i0:
            return np.arange(i0, i1 + 1, dtype=int)
        return np.r_[np.arange(i0, n, dtype=int), np.arange(0, i1 + 1, dtype=int)]
    else:
        if i0 >= i1:
            return np.arange(i0, i1 - 1, -1, dtype=int)
        return np.r_[np.arange(i0, -1, -1, dtype=int), np.arange(n - 1, i1 - 1, -1, dtype=int)]


def _sample_polyline_by_arclength(P, m):
    P = np.asarray(P, float)
    if P.shape[0] == 0:
        return P
    if P.shape[0] == 1:
        return np.repeat(P, m, axis=0)

    d = np.hypot(np.diff(P[:, 0]), np.diff(P[:, 1])) + 1e-12
    s = np.r_[0.0, np.cumsum(d)]
    L = float(s[-1])

    if L <= 1e-12:
        j = np.linspace(0, P.shape[0] - 1, m).round().astype(int)
        return P[j]

    t = np.linspace(0.0, L, m)
    x = np.interp(t, s, P[:, 0])
    y = np.interp(t, s, P[:, 1])
    return np.column_stack([x, y])


def _poly_contains(poly, pts):
    return bool(np.all(MplPath(poly).contains_points(pts, radius=1e-12)))


def _ensure_ccw(poly):
    p = np.asarray(poly, float)
    a = float(np.sum(p[:, 0] * np.roll(p[:, 1], -1) - np.roll(p[:, 0], -1) * p[:, 1]))
    if a < 0:
        p = p[::-1]
    return p


def calculate_curve_area(x, y, theta=None):
    """用 shoelace 公式对闭合采样曲线求面积（稳健，不依赖导数）"""
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if not (np.isclose(x[0], x[-1]) and np.isclose(y[0], y[-1])):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]
    area = 0.5 * np.abs(np.sum(x[:-1] * y[1:] - x[1:] * y[:-1]))
    return float(area)


def calculate_polygon_area(vertices):
    v = np.asarray(vertices, dtype=float)
    x = v[:, 0]
    y = v[:, 1]
    return 0.5 * np.abs(np.sum(x * np.roll(y, -1) - np.roll(x, -1) * y))

# =========================
# tight symmetric outer polygon (xy symmetry; concave allowed)
# =========================
def symmetric_outer_polygon_xy_concave(
    x, y, theta,
    num_vertices=24,
    delta_ratio=0.01,
    max_grow=80,
    bisect_iters=30,
    refine_rounds=6,
    refine_bisect=36,
):
    """
    xy 对称、允许凹的 tight outer polygon（尽可能小的 offset delta）
    追加：在外接约束下对部分顶点做内收，进一步减小面积。
    num_vertices 必须为 4 的倍数
    """
    if num_vertices % 4 != 0:
        raise ValueError("xy轴对称需要 num_vertices 为4的倍数")

    x = np.asarray(x, float)
    y = np.asarray(y, float)
    theta = np.asarray(theta, float)

    # close
    if not np.allclose([x[0], y[0]], [x[-1], y[-1]], atol=1e-12):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]
        theta = np.linspace(theta.min(), theta.max(), x.size, endpoint=False)

    # ensure CCW
    signed = float(np.sum(x[:-1] * y[1:] - x[1:] * y[:-1]))
    if signed < 0:
        x = x[::-1]
        y = y[::-1]
        theta = theta[::-1]

    # outward normals for CCW curve
    dx = np.gradient(x, theta)
    dy = np.gradient(y, theta)
    ds = np.hypot(dx, dy) + 1e-12
    nx = dy / ds
    ny = -dx / ds

    scale = float(np.hypot(x.max() - x.min(), y.max() - y.min()) + 1e-12)
    pts_curve = np.column_stack([x, y])

    q = num_vertices // 4
    mQ = q + 1
    tol = 1e-10 * scale

    def build_q1(delta):
        xo = x + delta * nx
        yo = y + delta * ny

        # +x and +y axis intersections
        pX, iX = _axis_crossings_positive(xo, yo, axis="x")
        pY, iY = _axis_crossings_positive(xo, yo, axis="y")

        n = xo.size

        seg_f = _wrap_indices(iX + 1, iY, n, forward=True)
        seg_b = _wrap_indices(iX + 1, iY, n, forward=False)

        def score(seg):
            if seg.size == 0:
                return -np.inf
            xs = xo[seg]
            ys = yo[seg]
            minxy = min(float(xs.min()), float(ys.min()))
            frac_q1 = float(np.mean((xs >= -tol) & (ys >= -tol)))
            return 1000.0 * frac_q1 + 10.0 * minxy - 1e-6 * seg.size

        seg = seg_f if score(seg_f) >= score(seg_b) else seg_b

        Pmid = np.column_stack([xo[seg], yo[seg]]) if seg.size else np.empty((0, 2), float)
        Praw = np.vstack([pX, Pmid, pY])

        Q1 = _sample_polyline_by_arclength(Praw, mQ)
        Q1[0] = pX
        Q1[-1] = pY
        Q1[0, 1] = 0.0
        Q1[-1, 0] = 0.0

        return Q1

    def build_poly_from_q1(Q1):
        Q2 = np.column_stack([-Q1[::-1, 0],  Q1[::-1, 1]])
        Q3 = -Q1
        Q4 = np.column_stack([ Q1[::-1, 0], -Q1[::-1, 1]])
        poly = np.vstack([Q1, Q2[1:], Q3[1:], Q4[1:-1]])
        return _ensure_ccw(poly)

    def _seg_intersect(a, b, c, d, eps=1e-12):
        def orient(p, q, r):
            return (q[0] - p[0]) * (r[1] - p[1]) - (q[1] - p[1]) * (r[0] - p[0])

        def onseg(p, q, r):
            return (
                min(p[0], r[0]) - eps <= q[0] <= max(p[0], r[0]) + eps
                and min(p[1], r[1]) - eps <= q[1] <= max(p[1], r[1]) + eps
            )

        o1 = orient(a, b, c)
        o2 = orient(a, b, d)
        o3 = orient(c, d, a)
        o4 = orient(c, d, b)

        if (o1 * o2 < -eps) and (o3 * o4 < -eps):
            return True
        if abs(o1) <= eps and onseg(a, c, b):
            return True
        if abs(o2) <= eps and onseg(a, d, b):
            return True
        if abs(o3) <= eps and onseg(c, a, d):
            return True
        if abs(o4) <= eps and onseg(c, b, d):
            return True
        return False

    def _is_simple_polygon(V):
        V = np.asarray(V, float)
        n = V.shape[0]
        if n < 4:
            return True
        for i in range(n):
            a = V[i]
            b = V[(i + 1) % n]
            for j in range(i + 1, n):
                if j == i:
                    continue
                if (j == i + 1) or ((i == 0) and (j == n - 1)):
                    continue
                if (i == j + 1) or ((j == 0) and (i == n - 1)):
                    continue
                c = V[j]
                d = V[(j + 1) % n]
                if _seg_intersect(a, b, c, d):
                    return False
        return True

    lo = 0.0
    hi = float(delta_ratio * scale)

    Q1 = build_q1(hi)
    poly = build_poly_from_q1(Q1)
    ok = _poly_contains(poly, pts_curve)

    grow = 0
    while (not ok) and grow < max_grow:
        lo = hi
        hi *= 1.35
        Q1 = build_q1(hi)
        poly = build_poly_from_q1(Q1)
        ok = _poly_contains(poly, pts_curve)
        grow += 1

    if not ok:
        return poly

    for _ in range(bisect_iters):
        mid = 0.5 * (lo + hi)
        Q1m = build_q1(mid)
        pm = build_poly_from_q1(Q1m)
        if _poly_contains(pm, pts_curve):
            hi = mid
            poly = pm
            Q1 = Q1m
        else:
            lo = mid

    if refine_rounds > 0:
        Q1 = _refine_q1_vertices(
            Q1,
            pts_curve,
            rounds=refine_rounds,
            bisect=refine_bisect,
            is_simple=_is_simple_polygon,
        )
        poly = build_poly_from_q1(Q1)

    return poly


def _refine_q1_vertices(Q1, pts_curve, rounds=2, bisect=22, is_simple=None):
    """在保持 xy 对称的前提下，让 Q1 顶点尽量靠近曲线。"""
    Q1 = np.asarray(Q1, float)
    pts = np.asarray(pts_curve, float)
    if Q1.shape[0] < 3:
        return Q1

    if is_simple is None:
        def is_simple(_):
            return True

    def build_poly(q1):
        q2 = np.column_stack([-q1[::-1, 0],  q1[::-1, 1]])
        q3 = -q1
        q4 = np.column_stack([ q1[::-1, 0], -q1[::-1, 1]])
        return np.vstack([q1, q2[1:], q3[1:], q4[1:-1]])

    def is_valid_q1(q1):
        poly_c = build_poly(q1)
        if not _poly_contains(poly_c, pts):
            return False
        if not is_simple(poly_c):
            return False
        return True

    poly = build_poly(Q1)
    if not _poly_contains(poly, pts):
        return Q1

    for _ in range(max(1, rounds)):
        moved = False
        for i in range(1, Q1.shape[0] - 1):
            v = Q1[i].copy()
            d2 = np.sum((pts - v) ** 2, axis=1)
            nearest = pts[int(np.argmin(d2))]
            direction = nearest - v
            if np.linalg.norm(direction) <= 1e-14:
                continue

            lo, hi = 0.0, 0.995
            best = v
            for _k in range(max(8, int(bisect))):
                mid = 0.5 * (lo + hi)
                cand = Q1.copy()
                cand[i] = v + mid * direction
                if is_valid_q1(cand):
                    lo = mid
                    best = cand[i]
                else:
                    hi = mid

            if np.linalg.norm(best - v) > 1e-9:
                Q1[i] = best
                moved = True

        # 端点沿轴线收紧（保持 x/y 轴对称）
        # Q1[0] 在 +x 轴上，Q1[-1] 在 +y 轴上
        def try_axis_shrink(idx, axis="x"):
            v = Q1[idx].copy()
            if axis == "x":
                hi = float(v[0])
                lo = 0.0
                if hi <= 1e-12:
                    return False
                best = v.copy()
                for _k in range(max(8, int(bisect))):
                    mid = 0.5 * (lo + hi)
                    cand = Q1.copy()
                    cand[idx] = np.array([mid, 0.0], float)
                    if is_valid_q1(cand):
                        best = cand[idx]
                        hi = mid
                    else:
                        lo = mid
                if abs(best[0] - v[0]) > 1e-9:
                    Q1[idx] = best
                    return True
                return False
            else:
                hi = float(v[1])
                lo = 0.0
                if hi <= 1e-12:
                    return False
                best = v.copy()
                for _k in range(max(8, int(bisect))):
                    mid = 0.5 * (lo + hi)
                    cand = Q1.copy()
                    cand[idx] = np.array([0.0, mid], float)
                    if is_valid_q1(cand):
                        best = cand[idx]
                        hi = mid
                    else:
                        lo = mid
                if abs(best[1] - v[1]) > 1e-9:
                    Q1[idx] = best
                    return True
                return False

        moved |= try_axis_shrink(0, axis="x")
        moved |= try_axis_shrink(Q1.shape[0] - 1, axis="y")

        # 边内缩：只处理 Q1 内部边，保持端点在轴上
        def try_edge_shrink(i0):
            if i0 <= 0 or i0 >= Q1.shape[0] - 2:
                return False
            p0 = Q1[i0].copy()
            p1 = Q1[i0 + 1].copy()
            e = p1 - p0
            L = float(np.hypot(e[0], e[1]))
            if L <= 1e-12:
                return False
            # 法线方向，选向“靠近原点”的方向作为内缩
            n = np.array([e[1], -e[0]], float) / L
            midp = 0.5 * (p0 + p1)
            if np.dot(n, midp) > 0.0:
                n = -n

            lo, hi = 0.0, 0.98
            best0, best1 = p0, p1
            for _k in range(max(8, int(bisect))):
                t = 0.5 * (lo + hi)
                cand = Q1.copy()
                cand[i0] = p0 + t * n
                cand[i0 + 1] = p1 + t * n
                # 仍需保持在第一象限
                if np.any(cand < -1e-10):
                    hi = t
                    continue
                if is_valid_q1(cand):
                    best0, best1 = cand[i0], cand[i0 + 1]
                    lo = t
                else:
                    hi = t

            if np.linalg.norm(best0 - p0) > 1e-9 or np.linalg.norm(best1 - p1) > 1e-9:
                Q1[i0] = best0
                Q1[i0 + 1] = best1
                return True
            return False

        for i0 in range(1, Q1.shape[0] - 2):
            moved |= try_edge_shrink(i0)

        if not moved:
            break

    return Q1


# =========================
# visualize (ONLY: xy symmetry / lobular2)
# =========================
def save_outer_polygon_to_excel(poly, area_poly, x, y, area_curve, filename):

    sheet_title = "outer_polygon_lobular2"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["X", "Y"])
    for vertex in poly:
        ws.append([float(vertex[0]), float(vertex[1])])

    ws2 = wb.create_sheet(title="curve_lobular2")
    ws2.append(["X", "Y"])
    for xi, yi in zip(x, y):
        ws2.append([float(xi), float(yi)])

    ws3 = wb.create_sheet(title="area_lobular2")
    ws3.append(["area_poly", "area_curve"])
    ws3.append([float(area_poly), float(area_curve)])

    wb.save(filename)


def visualize_lobular2(radius=4.17284, num_vertices=24, scale_ratio=1.0, output_xlsx='', plot: bool = True):

    plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False

    x, y, theta = create_lobular2(radius)

    # resample by arclength for smooth-ish plotting
    P = np.column_stack([x, y])
    P = np.vstack([P, P[0]])
    P_smooth = _sample_polyline_by_arclength(P, len(P) * 3)
    xs, ys = P_smooth[:, 0], P_smooth[:, 1]
    xs_scaled = xs * scale_ratio
    ys_scaled = ys * scale_ratio

    theta_smooth = np.linspace(0.0, 2 * np.pi, len(xs_scaled), endpoint=False)

    poly_scaled = symmetric_outer_polygon_xy_concave(
        xs_scaled,
        ys_scaled,
        theta_smooth,
        num_vertices=num_vertices,
        delta_ratio=0.01,
        refine_rounds=6,
        refine_bisect=36,
    )
    poly_closed = np.vstack([poly_scaled, poly_scaled[0]])

    area_curve = calculate_curve_area(x, y)
    area_poly = calculate_polygon_area(poly_scaled)

    if plot:
        plt.figure(figsize=(9, 7))
        plt.plot(xs, ys, "b-", lw=2, label=f"二叶草 (area={area_curve:.3f})")
        plt.plot(poly_closed[:, 0], poly_closed[:, 1], "r--", lw=2, label=f"outer polygon (area={area_poly:.3f}, n={num_vertices})")
        plt.plot(poly_scaled[:, 0], poly_scaled[:, 1], "ro", ms=5)

        plt.title(f"二叶草 & tight outer polygon (scaled area={area_poly})")
        plt.xlabel("X")
        plt.ylabel("Y")
        plt.grid(True, ls="--", alpha=0.5)
        plt.gca().set_aspect("equal", adjustable="box")
        plt.legend(loc="upper right")
        plt.show()

    save_outer_polygon_to_excel(poly_scaled, area_poly, x, y, area_curve, output_xlsx)
    return poly_scaled, area_poly, x, y, area_curve


def circum_lobular2(output_xlsx, scale_ratio_collision, num_vertices=16, plot: bool = True):
    return visualize_lobular2(
        radius=4.17284,
        num_vertices=num_vertices,
        scale_ratio=scale_ratio_collision,
        output_xlsx=str(output_xlsx),
        plot=plot,
    )

def main():

    scale_ratio_collision = 1.0
    # num_vertices 必须为 4 的倍数（xy 对称）
    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = data_dir / "outer_polygon_vertices_lobular2.xlsx"
    poly = circum_lobular2(output_xlsx=output_xlsx, scale_ratio_collision=scale_ratio_collision)

if __name__ == "__main__":
    main()
