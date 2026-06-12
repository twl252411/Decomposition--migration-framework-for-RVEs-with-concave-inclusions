# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from math import sqrt, pi, atan2

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from scipy.interpolate import make_interp_spline


# =========================
# 通用辅助函数
# =========================
def angle_from_points(center, point):
    """计算从中心点到目标点的角度(弧度)"""
    dx = point[0] - center[0]
    dy = point[1] - center[1]
    return atan2(dy, dx)


def generate_arc(center, start_angle, end_angle, radius, direction="COUNTERCLOCKWISE", num_points=100):
    """生成圆弧上的点集"""
    if direction == "COUNTERCLOCKWISE" and end_angle <= start_angle:
        end_angle += 2 * pi
    elif direction == "CLOCKWISE" and end_angle >= start_angle:
        end_angle -= 2 * pi

    angles = np.linspace(start_angle, end_angle, num_points)
    x = center[0] + radius * np.cos(angles)
    y = center[1] + radius * np.sin(angles)
    return np.column_stack((x, y))


# =========================
# 形状生成函数
# =========================
def _generate_lobular(arcs, num_points=100):
    """通用叶形生成函数"""
    all_points = []

    for arc in arcs:
        if "start" in arc and "end" in arc:
            start_angle = arc["start"]
            end_angle = arc["end"]
            radius = arc["radius"]
        else:
            center = arc["center"]
            point1 = arc["point1"]
            point2 = arc["point2"]
            start_angle = angle_from_points(center, point1)
            end_angle = angle_from_points(center, point2)
            radius = sqrt((point1[0] - center[0]) ** 2 + (point1[1] - center[1]) ** 2)

        arc_points = generate_arc(
            center=arc["center"],
            start_angle=start_angle,
            end_angle=end_angle,
            radius=radius,
            direction=arc["direction"],
            num_points=arc.get("num_points", num_points),
        )
        all_points.append(arc_points)

    combined = np.concatenate(all_points)
    x, y = combined[:, 0], combined[:, 1]

    # 这里只是给后续平滑/积分用的“参数”，不参与缩放
    theta = np.linspace(0, 2 * np.pi, len(x))
    return x, y, theta


def create_lobular2(radius=1.0):
    """生成二叶形曲线（二叶草）"""
    arcs = [
        # 左圆弧
        {"center": (-radius, 0.0), "start": pi / 3, "end": 5 * pi / 3, "radius": radius, "direction": "COUNTERCLOCKWISE"},
        # 下圆弧
        {"center": (0.0, -radius * sqrt(3)), "start": 2 * pi / 3, "end": pi / 3, "radius": radius, "direction": "CLOCKWISE"},
        # 右圆弧
        {"center": (radius, 0.0), "start": 4 * pi / 3, "end": 2 * pi / 3 + 2 * pi, "radius": radius, "direction": "COUNTERCLOCKWISE"},
        # 上圆弧
        {"center": (0.0, radius * sqrt(3)), "start": 5 * pi / 3, "end": 4 * pi / 3, "radius": radius, "direction": "CLOCKWISE", "num_points": 150},
    ]
    return _generate_lobular(arcs)


def create_kidney_shape(R1=12, R2=10, K1=15, a=1.5, num_points=200):
    """生成豌豆形（pea-like / kidney-like）曲线"""
    theta = np.linspace(0, 2 * pi, num_points)
    x = R1 * np.cos(theta) + K1 * np.exp(-a * np.cos(theta) - a)
    y = R2 * np.sin(theta)
    return x, y, theta


# =========================
# 面积计算（保留：用于打印信息/检查）
# =========================
def calculate_curve_area(x, y, theta):
    """使用格林公式计算曲线面积"""
    dx_dtheta = np.gradient(x, theta)
    dy_dtheta = np.gradient(y, theta)
    integrand = x * dy_dtheta - y * dx_dtheta
    integrate = np.trapezoid if hasattr(np, "trapezoid") else np.trapz
    return 0.5 * np.abs(integrate(integrand, theta))


def curve_area_shoelace(x, y):
    """使用鞋带公式计算闭合曲线面积"""
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if x.size < 3:
        return 0.0
    if not (np.isclose(x[0], x[-1]) and np.isclose(y[0], y[-1])):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]
    area = np.sum(x[:-1] * y[1:] - x[1:] * y[:-1])
    return 0.5 * np.abs(area)


def calculate_polygon_area(vertices):
    """计算多边形面积"""
    area = 0.0
    n = len(vertices)
    for i in range(n):
        j = (i + 1) % n
        area += vertices[i][0] * vertices[j][1] - vertices[j][0] * vertices[i][1]
    return 0.5 * np.abs(area)


# =========================
# 外接凹多边形（修复版）
# =========================
def create_concave_polygon(
    x,
    y,
    num_vertices=16,
    offset_tol=2e-4,
    max_iter=30,
    refine_rounds=2,
):
    """生成外接凹多边形：顶点到曲线距离尽量一致，面积尽量小。"""
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if x.size < 3:
        return np.vstack([np.column_stack([x, y]), [x[0], y[0]]]).T

    if not np.allclose([x[0], y[0]], [x[-1], y[-1]], atol=1e-9):
        x = np.append(x, x[0])
        y = np.append(y, y[0])

    x_open = x[:-1]
    y_open = y[:-1]
    if _calculate_signed_area(np.column_stack([x_open, y_open])) < 0:
        x_open = x_open[::-1]
        y_open = y_open[::-1]

    x_closed = np.r_[x_open, x_open[0]]
    y_closed = np.r_[y_open, y_open[0]]

    segx = np.diff(x_closed)
    segy = np.diff(y_closed)
    seglen = np.hypot(segx, segy)
    total_len = float(np.sum(seglen))
    if total_len <= 1e-12:
        return np.vstack([np.column_stack([x_open, y_open]), [x_open[0], y_open[0]]]).T

    cum_length = np.concatenate([[0.0], np.cumsum(seglen)])
    targets = np.linspace(0.0, total_len, num_vertices, endpoint=False)

    check_points = np.column_stack((x_open, y_open))
    from matplotlib.path import Path

    def build_polygon(offset: float) -> np.ndarray:
        verts = []
        for s in targets:
            i = int(np.searchsorted(cum_length, s, side="right") - 1)
            i = int(np.clip(i, 0, len(seglen) - 1))
            t = (s - cum_length[i]) / (seglen[i] + 1e-12)
            px = (1.0 - t) * x_closed[i] + t * x_closed[i + 1]
            py = (1.0 - t) * y_closed[i] + t * y_closed[i + 1]

            tx = x_closed[i + 1] - x_closed[i]
            ty = y_closed[i + 1] - y_closed[i]
            L = float(np.hypot(tx, ty))
            if L <= 1e-12:
                j = (i - 1) % len(seglen)
                tx = x_closed[j + 1] - x_closed[j]
                ty = y_closed[j + 1] - y_closed[j]
                L = float(np.hypot(tx, ty))
                if L <= 1e-12:
                    tx, ty, L = 1.0, 0.0, 1.0

            # 曲线为 CCW，外法线 = (ty, -tx)
            nx = ty / L
            ny = -tx / L
            verts.append([px + offset * nx, py + offset * ny])

        poly = np.asarray(verts, float)
        if _calculate_signed_area(poly) < 0:
            poly = poly[::-1]
        return poly

    def is_valid(poly: np.ndarray) -> bool:
        if poly.shape[0] < 3:
            return False
        if _calculate_signed_area(poly) <= 0:
            return False
        if not _is_simple_polygon(poly):
            return False
        return np.all(Path(poly).contains_points(check_points, radius=1e-10))

    scale = float(np.hypot(np.ptp(x_open), np.ptp(y_open)) + 1e-12)
    tol = float(offset_tol) * scale

    d_hi = max(scale * 0.002, tol)
    best = None
    for _ in range(30):
        poly = build_polygon(d_hi)
        if is_valid(poly):
            best = poly
            break
        d_hi *= 1.6

    if best is None:
        poly = build_polygon(d_hi)
        polygon_closed = np.vstack([poly, poly[0]])
        return polygon_closed.T

    d_lo = 0.0
    for _ in range(max_iter):
        d_mid = 0.5 * (d_lo + d_hi)
        poly = build_polygon(d_mid)
        if is_valid(poly):
            best = poly
            d_hi = d_mid
        else:
            d_lo = d_mid
        if (d_hi - d_lo) <= tol:
            break

    if refine_rounds > 0:
        best = _pull_vertices_toward_curve(best, check_points, rounds=refine_rounds)

    polygon_closed = np.vstack([best, best[0]])
    return polygon_closed.T


# =========================
# 外接凹多边形（lobular2：采用 circum_concave_general_1 的方法）
# =========================
def create_concave_polygon_lobular2(
    x,
    y,
    num_vertices=10,
    max_iterations=3,
    curvature_threshold=1.5,
    refine_rounds=3,
):
    """生成完全包裹曲线的外接凹多边形（lobular2：不强制对称，允许局部收紧）"""

    # 确保曲线闭合
    if not np.allclose([x[0], y[0]], [x[-1], y[-1]], atol=1e-9):
        x = np.append(x, x[0])
        y = np.append(y, y[0])

    # 曲率用于“加点”
    dx, dy = np.gradient(x), np.gradient(y)
    d2x, d2y = np.gradient(dx), np.gradient(dy)
    curvature = np.abs(dx * d2y - dy * d2x) / (dx**2 + dy**2 + 1e-12) ** 1.5

    # 用相邻点线段长度定义弧长
    segx = np.diff(x)
    segy = np.diff(y)
    seglen = np.hypot(segx, segy)
    curve_length = np.sum(seglen)
    cum_length = np.concatenate([[0.0], np.cumsum(seglen)])  # len = N

    # 1) 基础顶点：按弧长均匀取点（线性插值相邻点）
    base_vertices = []
    targets = np.linspace(0, curve_length, num_vertices, endpoint=False)
    for s in targets:
        i = np.searchsorted(cum_length, s, side="right") - 1
        i = int(np.clip(i, 0, len(seglen) - 1))
        ds = seglen[i] + 1e-12
        t = (s - cum_length[i]) / ds
        px = (1.0 - t) * x[i] + t * x[i + 1]
        py = (1.0 - t) * y[i] + t * y[i + 1]
        base_vertices.append((px, py))

    # 2) 高曲率区域加点
    extra_vertices = []
    high_curv = curvature > curvature_threshold
    step = max(1, len(high_curv) // max(1, (num_vertices // 2)))
    min_sep = 0.06 * curve_length / max(1, num_vertices)

    for i in range(0, len(high_curv), step):
        if high_curv[i] and 0 < i < len(x) - 1:
            dists = [np.hypot(x[i] - vx, y[i] - vy) for vx, vy in (base_vertices + extra_vertices)]
            if (len(dists) == 0) or (min(dists) > min_sep):
                extra_vertices.append((x[i], y[i]))

    # 合并、去重并排序
    vertices = np.unique(np.array(base_vertices + extra_vertices), axis=0)
    vertices = sorted(vertices.tolist(), key=lambda v: _get_arc_length(v, x, y, cum_length))

    # 确保逆时针
    if _calculate_signed_area(vertices) < 0:
        vertices = vertices[::-1]

    polygon = np.array(vertices, dtype=float)
    check_points = np.column_stack((x, y))

    from matplotlib.path import Path

    def contains_all(poly):
        return np.all(Path(poly).contains_points(check_points, radius=1e-10))

    # 3) 迭代外扩，确保包含全部点（径向外扩）
    for _ in range(max_iterations):
        if contains_all(polygon):
            break

        P = Path(polygon)
        inside = P.contains_points(check_points, radius=1e-10)
        outside = check_points[~inside]
        if outside.size == 0:
            break

        center = np.mean(polygon, axis=0)

        def min_dist_to_edges(p):
            return min(
                _point_to_segment_distance(p, polygon[i], polygon[(i + 1) % len(polygon)])
                for i in range(len(polygon))
            )

        max_dist = max(min_dist_to_edges(p) for p in outside)

        expand = 1.5 * max_dist + 1e-6
        vec = polygon - center
        norm = np.linalg.norm(vec, axis=1, keepdims=True) + 1e-12
        polygon = polygon + expand * (vec / norm)

        if _calculate_signed_area(polygon) < 0:
            polygon = polygon[::-1]

    if refine_rounds > 0:
        polygon = _pull_vertices_toward_curve(polygon, check_points, rounds=refine_rounds)

    polygon_closed = np.vstack([polygon, polygon[0]])
    return polygon_closed.T  # (poly_x, poly_y)

def _calculate_signed_area(polygon):
    """计算多边形的有向面积"""
    area = 0.0
    n = len(polygon)
    for i in range(n):
        j = (i + 1) % n
        area += polygon[i][0] * polygon[j][1] - polygon[j][0] * polygon[i][1]
    return area


def _point_to_segment_distance(p, a, b):
    """点 p 到线段 ab 的最短距离"""
    p = np.asarray(p, float)
    a = np.asarray(a, float)
    b = np.asarray(b, float)
    ab = b - a
    ap = p - a

    if np.dot(ab, ap) <= 0:
        return np.linalg.norm(ap)

    bp = p - b
    if np.dot(-ab, bp) <= 0:
        return np.linalg.norm(bp)

    return np.abs(ab[0] * ap[1] - ab[1] * ap[0]) / (np.linalg.norm(ab) + 1e-12)


def _get_arc_length(point, x, y, cum_length):
    """计算点在曲线上的近似弧长位置（用最近点索引）"""
    d = np.hypot(x - point[0], y - point[1])
    i = int(np.argmin(d))
    return cum_length[min(i, len(cum_length) - 1)]


# =========================
# 几何工具
# =========================


def _segment_intersect(p1, p2, q1, q2, eps=1e-12) -> bool:
    p1 = np.asarray(p1, float)
    p2 = np.asarray(p2, float)
    q1 = np.asarray(q1, float)
    q2 = np.asarray(q2, float)

    def orient(a, b, c):
        return float((b[0] - a[0]) * (c[1] - a[1]) - (b[1] - a[1]) * (c[0] - a[0]))

    def on_seg(a, b, c):
        return (
            min(a[0], b[0]) - eps <= c[0] <= max(a[0], b[0]) + eps
            and min(a[1], b[1]) - eps <= c[1] <= max(a[1], b[1]) + eps
        )

    o1 = orient(p1, p2, q1)
    o2 = orient(p1, p2, q2)
    o3 = orient(q1, q2, p1)
    o4 = orient(q1, q2, p2)

    if ((o1 > eps and o2 < -eps) or (o1 < -eps and o2 > eps)) and ((o3 > eps and o4 < -eps) or (o3 < -eps and o4 > eps)):
        return True

    if abs(o1) <= eps and on_seg(p1, p2, q1):
        return True
    if abs(o2) <= eps and on_seg(p1, p2, q2):
        return True
    if abs(o3) <= eps and on_seg(q1, q2, p1):
        return True
    if abs(o4) <= eps and on_seg(q1, q2, p2):
        return True
    return False


def _is_simple_polygon(poly: np.ndarray) -> bool:
    """判断多边形是否自交（poly 为 open polygon：N×2，不含重复闭合点）"""
    n = poly.shape[0]
    if n < 4:
        return True
    for i in range(n):
        a1 = poly[i]
        a2 = poly[(i + 1) % n]
        for j in range(i + 1, n):
            # 跳过相邻边与同一条边
            if j == i:
                continue
            if (j + 1) % n == i:
                continue
            if (i + 1) % n == j:
                continue

            b1 = poly[j]
            b2 = poly[(j + 1) % n]
            if _segment_intersect(a1, a2, b1, b2):
                return False
    return True


def _pull_vertices_toward_curve(polygon, check_points, rounds=2, contain_radius=1e-10):
    """在保持外接与简单性的前提下，让部分顶点靠近曲线以减小面积。"""
    from matplotlib.path import Path

    poly = np.array(polygon, dtype=float)
    curve = np.asarray(check_points, dtype=float)
    if len(poly) < 3 or curve.size == 0:
        return poly

    if _calculate_signed_area(poly) < 0:
        poly = poly[::-1]

    def contains_all(poly_arr):
        return np.all(Path(poly_arr).contains_points(curve, radius=contain_radius))

    if not contains_all(poly):
        return poly

    n = len(poly)
    for _ in range(max(1, rounds)):
        moved_any = False
        for i in range(n):
            v = poly[i].copy()
            d2 = np.sum((curve - v) ** 2, axis=1)
            nearest = curve[int(np.argmin(d2))]
            direction = nearest - v

            if np.linalg.norm(direction) <= 1e-14:
                continue

            lo, hi = 0.0, 0.98
            for _k in range(22):
                mid = 0.5 * (lo + hi)
                cand = poly.copy()
                cand[i] = v + mid * direction
                if _calculate_signed_area(cand) > 0 and _is_simple_polygon(cand) and contains_all(cand):
                    lo = mid
                else:
                    hi = mid

            if lo > 1e-9:
                poly[i] = v + lo * direction
                moved_any = True

        if not moved_any:
            break

    if _calculate_signed_area(poly) < 0:
        poly = poly[::-1]
    return poly


# =========================
# Excel 输出（保持你原来的 sheet 结构）
# =========================
def _sheet_titles(shape_name: str):
    if shape_name == "pea":
        return "outer_polygon_pea", "curve_pea", "area_pea"
    if shape_name == "lobular2":
        return "outer_polygon_lobular2", "curve_lobular2", "area_lobular2"
    raise ValueError(f"Unsupported shape name: {shape_name}")


def save_outer_polygon_to_excel(poly, area_poly, x, y, area_curve, filename, shape_name: str):
    sheet_title, curve_title, area_title = _sheet_titles(shape_name)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["X", "Y"])
    for vertex in poly:
        ws.append([float(vertex[0]), float(vertex[1])])

    ws2 = wb.create_sheet(title=curve_title)
    ws2.append(["X", "Y"])
    for xi, yi in zip(x, y):
        ws2.append([float(xi), float(yi)])

    ws3 = wb.create_sheet(title=area_title)
    ws3.append(["area_poly", "area_curve"])
    ws3.append([float(area_poly), float(area_curve)])

    wb.save(filename)


# =========================
# 可视化
# =========================
def visualize_shape(
    x,
    y,
    theta,
    curve_area,
    shape_name="Shape",
    num_vertices=16,
    smooth=True,
    smooth_factor=2,
    grid_size=1.0,
    x_margin_ratio=0.25,
    y_margin_ratio=0.25,
    scale_ratio=1.0,
    output_xlsx="",
    sheet_shape="pea",
    plot: bool = True,
):
    """
    可视化：曲线 + 外接凹多边形
    """
    # 可选平滑（只改变绘图与多边形拟合精度，不做尺度变换）
    if smooth and len(x) >= 6:
        theta_smooth = np.linspace(theta.min(), theta.max(), len(x) * smooth_factor)
        x_use = make_interp_spline(theta, x, k=2)(theta_smooth)
        y_use = make_interp_spline(theta, y, k=2)(theta_smooth)
        theta_use = theta_smooth
    else:
        x_use, y_use, theta_use = x, y, theta

    # 生成外接凹多边形
    if sheet_shape == "lobular2":
        # lobular2：沿用 circum_concave_general_1 的策略（不做缩放/收紧）
        poly_x, poly_y = create_concave_polygon_lobular2(
            x_use,
            y_use,
            num_vertices=num_vertices,
            refine_rounds=4,
        )
    else:
        # pea：统一偏移的外接多边形（顶点距离更均匀）
        x_scaled = x_use * scale_ratio
        y_scaled = y_use * scale_ratio
        poly_x, poly_y = create_concave_polygon(
            x_scaled,
            y_scaled,
            num_vertices=num_vertices,
        )
    polygon_vertices = list(zip(poly_x[:-1], poly_y[:-1]))
    poly_area = calculate_polygon_area(polygon_vertices)

    # 这里不输出多边形坐标/面积（如需调试，请在调用处自行打印）

    if plot:
        # 绘图
        plt.figure(figsize=(10, 8))
        plt.plot(x_use, y_use, "b-", linewidth=2, label=f"{shape_name} (area≈{curve_area:.4f})")
        plt.plot(poly_x, poly_y, "r--", linewidth=2, label=f"外接凹多边形 (area≈{poly_area:.4f})")
        plt.plot([v[0] for v in polygon_vertices], [v[1] for v in polygon_vertices], "ro", markersize=6)

        plt.title(f"{shape_name} 及其外接凹多边形（更贴合）", fontsize=16)
        plt.xlabel("X")
        plt.ylabel("Y")

        # 留白 + 网格
        x_min, x_max = np.min(x_use), np.max(x_use)
        y_min, y_max = np.min(y_use), np.max(y_use)

        x_pad = x_margin_ratio * (x_max - x_min + 1e-12)
        y_pad = y_margin_ratio * (y_max - y_min + 1e-12)

        plt.xlim(x_min - x_pad, x_max + x_pad)
        plt.ylim(y_min - y_pad, y_max + y_pad)

        ax = plt.gca()
        ax.set_aspect("equal", adjustable="box")

        ax.set_xticks(np.arange(np.floor(x_min - x_pad), np.ceil(x_max + x_pad) + 1e-9, grid_size))
        ax.set_yticks(np.arange(np.floor(y_min - y_pad), np.ceil(y_max + y_pad) + 1e-9, grid_size))
        ax.grid(True, linestyle="--", alpha=0.6)

        plt.legend(fontsize=12, loc="upper right")
        plt.show()

    if output_xlsx:
        save_outer_polygon_to_excel(
            polygon_vertices,
            poly_area,
            x,
            y,
            curve_area,
            output_xlsx,
            shape_name=sheet_shape,
        )

    return polygon_vertices


# =========================
# 主函数：只保留 2 个选项
# =========================
def x_axis_intersection_center(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if x.size == 0:
        raise ValueError("Curve vertices are empty.")
    if x.size < 2:
        return float(x[0]), 0.0
    if not (np.isclose(x[0], x[-1]) and np.isclose(y[0], y[-1])):
        x = np.r_[x, x[0]]
        y = np.r_[y, y[0]]

    xs = []
    for i in range(x.size - 1):
        x1, y1 = float(x[i]), float(y[i])
        x2, y2 = float(x[i + 1]), float(y[i + 1])
        if np.isclose(y1, 0.0):
            xs.append(x1)
        if np.isclose(y2, 0.0):
            xs.append(x2)
        if (y1 > 0.0 and y2 < 0.0) or (y1 < 0.0 and y2 > 0.0):
            t = -y1 / (y2 - y1)
            xs.append(x1 + t * (x2 - x1))

    if not xs:
        return float(np.mean(x[:-1])), float(np.mean(y[:-1]))

    x_min = min(xs)
    x_max = max(xs)
    return 0.5 * (x_min + x_max), 0.0


def visualize_concave_general(
    shape_name="pea",
    scale_ratio=1.0,
    num_vertices=20,      # pea 默认 20
    num_points=240,
    output_xlsx="",
    plot: bool = True,
):
    plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False

    if shape_name == "pea":
        x, y, theta = create_kidney_shape(R1=9.84, R2=8.2, K1=12.3, a=1.4, num_points=num_points)
        center_x, center_y = x_axis_intersection_center(x, y)
        x = (np.asarray(x, float) - center_x)
        y = (np.asarray(y, float) - center_y)
        curve_area = curve_area_shoelace(x, y)
        grid_size = 2.0

    elif shape_name == "lobular2":
        x, y, theta = create_lobular2(radius=4.17284)
        curve_area = calculate_curve_area(x, y, theta)
        grid_size = 0.5

    else:
        raise ValueError("shape_name must be 'pea' or 'lobular2'")

    polygon_vertices = visualize_shape(
        x,
        y,
        theta,
        curve_area=curve_area,
        shape_name=shape_name,
        num_vertices=num_vertices,
        grid_size=grid_size,
        scale_ratio=scale_ratio,
        output_xlsx=output_xlsx,
        sheet_shape=shape_name,
        plot=plot,
    )

    poly_area = calculate_polygon_area(polygon_vertices)
    return polygon_vertices, poly_area, x, y, curve_area


def circum_concave_general(
    output_xlsx,
    scale_ratio_collision=1.025,
    shape_name="pea",
    plot: bool = True,
    num_vertices: int = 16,
    num_points: int = 240,
):
    return visualize_concave_general(
        shape_name=shape_name,
        scale_ratio=scale_ratio_collision,
        num_vertices=num_vertices,
        num_points=num_points,
        output_xlsx=str(output_xlsx),
        plot=plot,
    )

def main():
    scale_ratio_collision = 1.00
    shape_name = ["lobular2", "pea"][1]

    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)

    output_xlsx = data_dir / f"outer_polygon_vertices_{shape_name}.xlsx"

    circum_concave_general(
        output_xlsx=output_xlsx,
        scale_ratio_collision=scale_ratio_collision,
        shape_name=shape_name,
    )


if __name__ == "__main__":
    main()
