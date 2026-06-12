# file: convex_partition_pea_final.py
# -*- coding: utf-8 -*-

import math
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

PRECISION = 6

Point = Tuple[float, float]
Polygon = List[Point]

class GeometryUtils:

    @staticmethod
    def calculate_signed_area(polygon: Polygon) -> float:
        area = 0.0
        n = len(polygon)
        for i in range(n):
            j = (i + 1) % n
            area += polygon[i][0] * polygon[j][1] - polygon[j][0] * polygon[i][1]
        return area

    @staticmethod
    def cross(o: Point, a: Point, b: Point) -> float:
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    @staticmethod
    def distance(a: Point, b: Point) -> float:
        return math.hypot(a[0] - b[0], a[1] - b[1])

    @staticmethod
    def round_coord(coord):
        if isinstance(coord, (list, tuple)):
            return tuple(round(c, PRECISION) for c in coord)
        return round(coord, PRECISION)

    @staticmethod
    def scale_of_polygon(polygon: Polygon) -> float:
        xs = [p[0] for p in polygon]
        ys = [p[1] for p in polygon]
        dx = max(xs) - min(xs)
        dy = max(ys) - min(ys)
        return max(dx, dy, 1.0)


class PolygonUtils:
    @staticmethod
    def ensure_ccw(polygon: Polygon, epsilon: float = 1e-12) -> Polygon:
        area = GeometryUtils.calculate_signed_area(polygon)
        if abs(area) < epsilon:
            raise ValueError(f"Degenerate polygon detected (signed area ~ 0): {area:.3e}")
        return polygon if area > 0 else list(reversed(polygon))

    @staticmethod
    def is_convex(polygon: Polygon, eps: float = 1e-12) -> bool:
        poly = PolygonUtils.ensure_ccw(polygon)
        n = len(poly)
        if n < 3:
            return False
        s = GeometryUtils.scale_of_polygon(poly)
        tol = max(eps, (10 ** (-PRECISION)) * s * s * 5.0)
        for i in range(n):
            a = poly[i]
            b = poly[(i + 1) % n]
            c = poly[(i + 2) % n]
            if GeometryUtils.cross(a, b, c) < -tol:
                return False
        return True

    @staticmethod
    def is_point_inside_polygon(p: Point, polygon: Polygon) -> bool:
        x, y = p
        n = len(polygon)
        inside = False
        eps = 1e-12

        for i in range(n):
            j = (i + 1) % n
            xi, yi = polygon[i]
            xj, yj = polygon[j]

            cross_val = GeometryUtils.cross(polygon[i], polygon[j], p)
            if abs(cross_val) <= 1e-10:
                if (min(xi, xj) - 1e-10 <= x <= max(xi, xj) + 1e-10 and
                        min(yi, yj) - 1e-10 <= y <= max(yi, yj) + 1e-10):
                    return True

            intersects = ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi + eps) + xi)
            if intersects:
                inside = not inside

        return inside

    @staticmethod
    def is_simple_polygon(polygon: Polygon, timeout_seconds: float = 5.0) -> bool:
        import time

        start = time.time()
        poly = polygon
        n = len(poly)
        if n < 3:
            return False

        seen = set()
        for v in poly:
            rv = GeometryUtils.round_coord(v)
            if rv in seen:
                return False
            seen.add(rv)

        for i in range(n):
            if time.time() - start > timeout_seconds:
                return False
            a = poly[i]
            b = poly[(i + 1) % n]
            for j in range(i + 2, n):
                if time.time() - start > timeout_seconds:
                    return False
                c = poly[j]
                d = poly[(j + 1) % n]
                if (j + 1) % n == i:
                    continue
                if SegmentUtils.segment_intersect(a, b, c, d):
                    return False
        return True


class SegmentUtils:
    @staticmethod
    def segment_intersect(a: Point, b: Point, c: Point, d: Point, epsilon: float = 1e-10) -> bool:
        def ccw(A, B, C) -> float:
            v = (B[0] - A[0]) * (C[1] - A[1]) - (B[1] - A[1]) * (C[0] - A[0])
            return 0.0 if abs(v) <= epsilon else v

        def on_segment(p, q, r) -> bool:
            return (min(p[0], r[0]) - epsilon <= q[0] <= max(p[0], r[0]) + epsilon and
                    min(p[1], r[1]) - epsilon <= q[1] <= max(p[1], r[1]) + epsilon and
                    ccw(p, q, r) == 0.0)

        ab_c = ccw(a, b, c)
        ab_d = ccw(a, b, d)
        cd_a = ccw(c, d, a)
        cd_b = ccw(c, d, b)

        if (ab_c * ab_d < 0.0) and (cd_a * cd_b < 0.0):
            return True

        if on_segment(a, c, b) or on_segment(a, d, b) or on_segment(c, a, d) or on_segment(c, b, d):
            return True

        return False


@dataclass(frozen=True)
class PartitionParams:
    merge_tip_allow_slanted: bool = True
    merge_tip_y_ratio: float = 0.70
    fan_count: int = 19
    merge_enable: bool = True
    merge_horiz_dy_eps: float = 0.40
    merge_max_iter: int = 80
    plot: bool = True
    sheet_name: str = "Outer_Polygon"
    simplify_collinear_eps: float = 1e-10
    visibility_mid_samples: int = 3
    visibility_max_candidates: int = 32


def chain_ccw_indices(n: int, i0: int, i1: int) -> List[int]:
    out = [i0]
    i = i0
    while i != i1:
        i = (i + 1) % n
        out.append(i)
        if len(out) > n + 5:
            break
    return out


def chain_cw_indices(n: int, i0: int, i1: int) -> List[int]:
    out = [i0]
    i = i0
    while i != i1:
        i = (i - 1) % n
        out.append(i)
        if len(out) > n + 5:
            break
    return out


def chain_from_to_in_list(chain: List[int], start: int, end: int, forward: bool = True) -> List[int]:
    pos = {v: i for i, v in enumerate(chain)}
    if start not in pos or end not in pos:
        return []
    i0 = pos[start]
    i1 = pos[end]
    if forward:
        if i0 <= i1:
            return chain[i0:i1 + 1]
        return chain[i0:] + chain[:i1 + 1]
    else:
        if i1 <= i0:
            sub = chain[i1:i0 + 1]
            sub.reverse()
            return sub
        sub = chain[:i0 + 1] + chain[i1:]
        sub.reverse()
        return sub


def chords_add_caps(
    chords_mid: List[Tuple[int, int]],
    itop: int,
    ibot: int,
    rtop: int,
    rbot: int,
    right_order: Dict[int, int],
) -> List[Tuple[int, int]]:
    top = (itop, rtop)
    bot = (ibot, rbot)

    seq = [top] + chords_mid + [bot]

    out: List[Tuple[int, int]] = []
    seen = set()
    last_ord = -10**9
    for ip, iq in seq:
        if (ip, iq) in seen:
            continue
        ordq = right_order.get(iq, None)
        if ordq is None:
            continue
        if ordq <= last_ord:
            continue
        out.append((ip, iq))
        seen.add((ip, iq))
        last_ord = ordq
    return out


def convex_partition_pea_vertex_fan_chain_dp(polygon: Polygon, params: PartitionParams) -> List[Polygon]:
    poly0 = PolygonUtils.ensure_ccw(simplify_polygon_collinear(polygon, eps=params.simplify_collinear_eps))
    if not PolygonUtils.is_simple_polygon(poly0):
        raise ValueError("Input polygon is not simple (self-intersecting).")

    n = len(poly0)
    itop = max(range(n), key=lambda i: (poly0[i][1], -poly0[i][0]))
    ibot = min(range(n), key=lambda i: (poly0[i][1], poly0[i][0]))

    ccw = chain_ccw_indices(n, itop, ibot)
    cw = chain_cw_indices(n, itop, ibot)

    meanx_ccw = sum(poly0[i][0] for i in ccw) / max(1, len(ccw))
    meanx_cw = sum(poly0[i][0] for i in cw) / max(1, len(cw))

    left_chain = ccw if meanx_ccw <= meanx_cw else cw
    right_chain = cw if left_chain is ccw else ccw

    right_order: Dict[int, int] = {vid: k for k, vid in enumerate(right_chain)}

    rtop = right_chain[0]
    rbot = right_chain[-1]

    anchors = select_fan_anchor_vertices(poly0, left_chain, fan_count=params.fan_count)
    anchors = sorted(set(anchors + [itop, ibot]), key=lambda i: (-poly0[i][1], poly0[i][0]))

    candidates = visible_target_candidates(poly0, anchors, right_chain, right_order, params)
    chords_mid = select_monotone_chords_dp(anchors, candidates)

    chords = chords_add_caps(chords_mid, itop, ibot, rtop, rbot, right_order)
    if len(chords) < 2:
        return [poly0]

    parts: List[Polygon] = []
    for k in range(len(chords) - 1):
        ip0, iq0 = chords[k]
        ip1, iq1 = chords[k + 1]

        left_path = chain_from_to_in_list(left_chain, ip0, ip1, forward=True)
        right_path = chain_from_to_in_list(right_chain, iq1, iq0, forward=False)  # FIX: right boundary must go upward

        if not left_path or not right_path:
            continue

        cell = [poly0[i] for i in left_path]
        tail = [poly0[i] for i in right_path]
        if tail and GeometryUtils.round_coord(cell[-1]) == GeometryUtils.round_coord(tail[0]):
            cell.extend(tail[1:])
        else:
            cell.extend(tail)

        eps_merge = 10 ** (-PRECISION)

        cell = polygon_clean(cell, eps=eps_merge)
        cell = simplify_polygon_collinear(cell, eps=params.simplify_collinear_eps)
        cell = polygon_clean(cell, eps=eps_merge)

        if len(cell) < 3:
            continue

        area = GeometryUtils.calculate_signed_area(cell)
        if abs(area) <= 1e-14:
            continue

        cell = cell if area > 0 else list(reversed(cell))

        if not PolygonUtils.is_simple_polygon(cell, timeout_seconds=1.5):
            continue
        if not PolygonUtils.is_convex(cell):
            continue
        parts.append(cell)

    parts = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=params.simplify_collinear_eps)) for p in parts]
    parts = [p for p in parts if len(p) >= 3 and PolygonUtils.is_convex(p)]
    parts.sort(key=lambda p: -abs(GeometryUtils.calculate_signed_area(p)))
    return parts if parts else [poly0]


def edge_canon(a: Point, b: Point) -> Tuple[Point, Point]:
    ra = GeometryUtils.round_coord(a)
    rb = GeometryUtils.round_coord(b)
    return (ra, rb) if ra <= rb else (rb, ra)


def edge_is_horizontal(a: Point, b: Point, dy_eps: float) -> bool:
    return abs(a[1] - b[1]) <= dy_eps


def edge_map_build(polys: List[Polygon]) -> Dict[Tuple[Point, Point], List[int]]:
    m: Dict[Tuple[Point, Point], List[int]] = {}
    for pid, p in enumerate(polys):
        n = len(p)
        for i in range(n):
            e = edge_canon(p[i], p[(i + 1) % n])
            m.setdefault(e, []).append(pid)
    return m


def load_polygon_vertices_from_excel(xlsx_path: str, sheet_name: str = "Outer_Polygon") -> Polygon:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    vertices: Polygon = []
    for row in rows[1:]:
        if row is None or len(row) < 2 or row[0] is None or row[1] is None:
            break
        vertices.append((float(row[0]), float(row[1])))

    if len(vertices) < 3:
        raise ValueError("Polygon vertex count < 3 in Excel.")
    return vertices


def merge_two_convex(poly1: Polygon, poly2: Polygon, eps: float, simplify_eps: float) -> Optional[Polygon]:
    def same_pt(u: Point, v: Point) -> bool:
        return abs(u[0] - v[0]) < eps and abs(u[1] - v[1]) < eps

    p1 = PolygonUtils.ensure_ccw(poly1)
    p2 = PolygonUtils.ensure_ccw(poly2)
    n1, n2 = len(p1), len(p2)
    if n1 < 3 or n2 < 3:
        return None

    shared = None
    for i in range(n1):
        a1 = p1[i]
        b1 = p1[(i + 1) % n1]
        for j in range(n2):
            a2 = p2[j]
            b2 = p2[(j + 1) % n2]
            if same_pt(a1, b2) and same_pt(b1, a2):
                shared = (i, j, "rev")
                break
            if same_pt(a1, a2) and same_pt(b1, b2):
                shared = (i, j, "same")
                break
        if shared is not None:
            break

    if shared is None:
        return None

    i, j, mode = shared
    if mode == "rev":
        ja = (j + 1) % n2
        jb = j
    else:
        ja = j
        jb = (j + 1) % n2

    merged: List[Point] = []

    k = (i + 1) % n1
    merged.append(p1[k])
    while k != i:
        k = (k + 1) % n1
        merged.append(p1[k])

    k = (ja + 1) % n2
    while k != jb:
        merged.append(p2[k])
        k = (k + 1) % n2

    cleaned: List[Point] = []
    for pt in merged:
        if not cleaned:
            cleaned.append(pt)
        else:
            if not same_pt(cleaned[-1], pt):
                cleaned.append(pt)

    if len(cleaned) >= 2 and same_pt(cleaned[0], cleaned[-1]):
        cleaned.pop()

    if len(cleaned) < 3:
        return None

    cleaned = simplify_polygon_collinear(cleaned, eps=simplify_eps)
    if len(cleaned) < 3:
        return None

    cleaned = PolygonUtils.ensure_ccw(cleaned)
    if not PolygonUtils.is_simple_polygon(cleaned, timeout_seconds=2.0):
        return None
    if not PolygonUtils.is_convex(cleaned):
        return None

    return cleaned


def mirror_signature(poly: Polygon, eps_round: int = PRECISION) -> Tuple[Tuple[float, float], ...]:
    pts = [(round(p[0], eps_round), round(p[1], eps_round)) for p in poly]
    pts.sort()
    return tuple(pts)


def mirror_signature_y(poly: Polygon, eps_round: int = PRECISION) -> Tuple[Tuple[float, float], ...]:
    pts = [(round(p[0], eps_round), round(-p[1], eps_round)) for p in poly]
    pts.sort()
    return tuple(pts)


def merge_horizontal_symmetric(parts: List[Polygon], params: PartitionParams) -> List[Polygon]:
    if not params.merge_enable:
        return parts

    eps = 10 ** (-PRECISION)
    dy_eps = float(params.merge_horiz_dy_eps)
    simplify_eps = float(params.simplify_collinear_eps)

    def centroid_y(p: Polygon) -> float:
        ys = [v[1] for v in p]
        return 0.5 * (min(ys) + max(ys))

    def build_sig_map(polys: List[Polygon]) -> Dict[Tuple[Tuple[float, float], ...], int]:
        sm: Dict[Tuple[Tuple[float, float], ...], int] = {}
        for i, p in enumerate(polys):
            sm[mirror_signature(p)] = i
        return sm

    it = 0
    polys = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=simplify_eps)) for p in parts]
    polys = [p for p in polys if len(p) >= 3 and PolygonUtils.is_convex(p)]
    polys.sort(key=lambda p: -abs(GeometryUtils.calculate_signed_area(p)))

    ymax_abs = max(abs(v[1]) for p in polys for v in p) if polys else 0.0
    tip_thr = float(params.merge_tip_y_ratio) * float(ymax_abs)

    while it < int(params.merge_max_iter):
        it += 1
        sig_map = build_sig_map(polys)
        e_map = edge_map_build(polys)

        best = None  # (priority, i, j, mi_i, mi_j, mup, mdn)

        for e, ids in e_map.items():
            if len(ids) != 2:
                continue
            i, j = ids[0], ids[1]
            if i == j:
                continue

            a, b = e[0], e[1]

            is_h = edge_is_horizontal(a, b, dy_eps=dy_eps)

            if not is_h:
                if not bool(getattr(params, "merge_tip_allow_slanted", True)):
                    continue

                cy_edge = 0.5 * (a[1] + b[1])
                if abs(cy_edge) <= tip_thr:
                    continue

            mi = merge_two_convex(polys[i], polys[j], eps=eps, simplify_eps=simplify_eps)
            if mi is None:
                continue

            cy = centroid_y(mi)
            area = abs(GeometryUtils.calculate_signed_area(mi)) * 0.5
            priority = (-abs(cy), -area)

            if cy > 1e-9:
                mi_i = sig_map.get(mirror_signature_y(polys[i]))
                mi_j = sig_map.get(mirror_signature_y(polys[j]))
                if mi_i is None or mi_j is None or mi_i == mi_j:
                    continue
                md = merge_two_convex(polys[mi_i], polys[mi_j], eps=eps, simplify_eps=simplify_eps)
                if md is None:
                    continue
                cand = (priority, i, j, mi_i, mi_j, mi, md)
            elif cy < -1e-9:
                continue
            else:
                cand = (priority, i, j, None, None, mi, None)

            if best is None or cand[0] < best[0]:
                best = cand

        if best is None:
            break

        _, i, j, mi_i, mi_j, mup, mdn = best

        keep: List[Polygon] = []
        for k, p in enumerate(polys):
            if k in (i, j):
                continue
            if mi_i is not None and k in (mi_i, mi_j):
                continue
            keep.append(p)

        keep.append(mup)
        if mdn is not None:
            keep.append(mdn)

        polys = [PolygonUtils.ensure_ccw(simplify_polygon_collinear(p, eps=simplify_eps)) for p in keep]
        polys = [p for p in polys if len(p) >= 3 and PolygonUtils.is_convex(p)]
        polys.sort(key=lambda p: -abs(GeometryUtils.calculate_signed_area(p)))

    return polys


def polygon_clean(polygon: Polygon, eps: float) -> Polygon:
    if not polygon:
        return []

    out: List[Point] = []
    for p in polygon:
        if not out:
            out.append(p)
            continue
        if abs(out[-1][0] - p[0]) <= eps and abs(out[-1][1] - p[1]) <= eps:
            continue
        out.append(p)

    if len(out) >= 2 and abs(out[0][0] - out[-1][0]) <= eps and abs(out[0][1] - out[-1][1]) <= eps:
        out.pop()

    seen = set()
    uniq: List[Point] = []
    for p in out:
        rp = GeometryUtils.round_coord(p)
        if rp in seen:
            continue
        seen.add(rp)
        uniq.append(p)

    return uniq


def plot_partition(original: Polygon, parts: List[Polygon]) -> None:
    plt.figure(figsize=(10, 8))

    ox, oy = zip(*original)
    plt.plot(ox + (ox[0],), oy + (oy[0],), "k--", lw=2, label="Outer Polygon")

    colors = plt.cm.tab20.colors

    for i, p in enumerate(parts):
        px, py = zip(*p)
        plt.fill(px, py, color=colors[i % len(colors)], alpha=0.25)
        plt.plot(px + (px[0],), py + (py[0],), "-", lw=2, color=colors[i % len(colors)])

    plt.title(f"Pea Vertex-Fan Chain-DP Convex Partition (parts={len(parts)})")
    plt.axis("equal")
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.show()


def _copy_sheet_with_format(src_wb, dest_wb, sheet_name: str) -> None:
    if sheet_name not in src_wb.sheetnames:
        return
    if sheet_name in dest_wb.sheetnames:
        del dest_wb[sheet_name]
    src_ws = src_wb[sheet_name]
    dest_ws = dest_wb.create_sheet(title=sheet_name)
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))


def save_partition_to_excel(
    output_xlsx: str,
    original: Polygon,
    parts: List[Polygon],
    params: PartitionParams,
    source_xlsx: str,
) -> None:

    wb = Workbook()
    ws = wb.active

    # 主表名：尽量沿用你现在的参数命名；没有就给一个默认值
    ws.title = str(getattr(params, "sheet_name", "outer_polygon_pea"))

    # 1) 写入 Original Polygon（与你的格式一致）
    ws.append(["Original Polygon X", "Original Polygon Y"])
    for x, y in original:
        ws.append([float(x), float(y)])

    # 2) 写入 Convex Parts（空行分段 + 指定表头）
    for i, part in enumerate(parts, start=1):
        ws.append([])
        ws.append([f"Convex Part {i} X", f"Convex Part {i} Y"])
        for x, y in part:
            ws.append([float(x), float(y)])

    # 3) Metadata sheet（与你的格式一致：Attribute / Value）
    meta = wb.create_sheet(title="Metadata")
    meta.append(["Attribute", "Value"])

    # 统计：凹点数量（对 original 做一个稳健计算）
    def _count_concave_vertices(poly: Polygon, eps: float = 1e-12) -> int:
        poly_ccw = PolygonUtils.ensure_ccw(poly)
        n = len(poly_ccw)
        if n < 3:
            return 0
        cnt = 0
        for k in range(n):
            a = poly_ccw[(k - 1) % n]
            b = poly_ccw[k]
            c = poly_ccw[(k + 1) % n]
            if GeometryUtils.cross(a, b, c) < -eps:
                cnt += 1
        return cnt

    # PCA AR 统计（你已有 polygon_aspect_ratio_pca）

    meta.append(["Concave Points", _count_concave_vertices(original)])
    meta.append(["Convex Parts", len(parts)])
    meta.append(["Original Vertices", len(original)])
    meta.append(["ArLimit", float(getattr(params, "ar_limit", 0.0))])
    meta.append(["SoftFallback", int(bool(getattr(params, "soft_fallback", False)))])

    source_wb = load_workbook(source_xlsx)
    _copy_sheet_with_format(source_wb, wb, "curve_pea")
    _copy_sheet_with_format(source_wb, wb, "area_pea")

    wb.save(output_xlsx)


def segment_is_visible_inside_polygon(poly: Polygon, p: Point, q: Point, mid_samples: int) -> bool:
    n = len(poly)
    eps = 1e-10
    rp = GeometryUtils.round_coord(p)
    rq = GeometryUtils.round_coord(q)

    for i in range(n):
        a = poly[i]
        b = poly[(i + 1) % n]
        ra = GeometryUtils.round_coord(a)
        rb = GeometryUtils.round_coord(b)

        if ra == rp or rb == rp or ra == rq or rb == rq:
            continue

        if SegmentUtils.segment_intersect(p, q, a, b, epsilon=eps):
            return False

    for k in range(1, mid_samples + 1):
        t = k / (mid_samples + 1)
        m = (p[0] * (1 - t) + q[0] * t, p[1] * (1 - t) + q[1] * t)
        if not PolygonUtils.is_point_inside_polygon(m, poly):
            return False

    return True


def select_fan_anchor_vertices(poly: Polygon, left_chain: List[int], fan_count: int) -> List[int]:
    ids = list(left_chain)
    ids.sort(key=lambda i: (-poly[i][1], poly[i][0]))
    if len(ids) < 3:
        return ids

    m = max(9, int(fan_count))
    if m > len(ids):
        return ids

    idxs = [round(k * (len(ids) - 1) / (m - 1)) for k in range(m)]
    picks = [ids[int(ii)] for ii in idxs]

    top = ids[0]
    bot = ids[-1]
    mid = min(ids, key=lambda i: abs(poly[i][1]))
    picks.extend([top, mid, bot])

    out = sorted(set(picks), key=lambda i: (-poly[i][1], poly[i][0]))
    return out


def select_monotone_chords_dp(
    anchors: List[int],
    candidates: Dict[int, List[Tuple[int, float, int]]],
) -> List[Tuple[int, int]]:
    NEG_INF = -10**18

    A = anchors
    m = len(A)

    states = []
    for ip in A:
        for iq, cost, order in candidates.get(ip, []):
            states.append((ip, cost, order))
    if not states:
        return []

    orders = sorted({order for _, _, order in states})
    order_to_idx = {o: i for i, o in enumerate(orders)}
    K = len(orders)

    dp = [[NEG_INF] * K for _ in range(m)]
    back: List[List[Optional[Tuple[int, int]]]] = [[None] * K for _ in range(m)]
    pick: List[List[Optional[int]]] = [[None] * K for _ in range(m)]

    cand_by_i: List[List[Tuple[int, float, int]]] = []
    for i in range(m):
        ip = A[i]
        cand_by_i.append(candidates.get(ip, []))

    for i in range(m):
        for iq, cost, order in cand_by_i[i]:
            k = order_to_idx[order]
            val = 1.0 - cost
            if val > dp[i][k]:
                dp[i][k] = val
                back[i][k] = None
                pick[i][k] = iq

    for i in range(1, m):
        for iq, cost, order in cand_by_i[i]:
            k = order_to_idx[order]
            best_prev = NEG_INF
            best_prev_k = None
            for kk in range(0, k):
                if dp[i - 1][kk] > best_prev:
                    best_prev = dp[i - 1][kk]
                    best_prev_k = kk
            if best_prev_k is not None:
                val = best_prev + (1.0 - cost)
                if val > dp[i][k]:
                    dp[i][k] = val
                    back[i][k] = (i - 1, best_prev_k)
                    pick[i][k] = iq

        for k in range(K):
            if dp[i - 1][k] > dp[i][k]:
                dp[i][k] = dp[i - 1][k]
                back[i][k] = (i - 1, k)
                pick[i][k] = pick[i - 1][k]

    best_val = NEG_INF
    best_pos = None
    for k in range(K):
        if dp[m - 1][k] > best_val:
            best_val = dp[m - 1][k]
            best_pos = (m - 1, k)

    if best_pos is None or best_val <= NEG_INF / 2:
        return []

    seq: List[Tuple[int, int]] = []
    cur = best_pos
    while cur is not None:
        i, k = cur
        ip = A[i]
        iq = pick[i][k]
        if iq is not None:
            seq.append((ip, iq))
        cur = back[i][k]

    seq.reverse()

    seq2: List[Tuple[int, int]] = []
    seen_iq = set()
    for ip, iq in seq:
        if iq in seen_iq:
            continue
        seen_iq.add(iq)
        seq2.append((ip, iq))

    return seq2


def simplify_polygon_collinear(polygon: Polygon, eps: float = 1e-10) -> Polygon:
    poly = list(polygon)
    if len(poly) < 3:
        return poly

    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        keep: List[Point] = []
        for i in range(n):
            a = poly[(i - 1) % n]
            b = poly[i]
            c = poly[(i + 1) % n]
            if GeometryUtils.distance(a, b) < eps:
                changed = True
                continue
            if abs(GeometryUtils.cross(a, b, c)) <= eps and (
                    min(a[0], c[0]) - eps <= b[0] <= max(a[0], c[0]) + eps and
                    min(a[1], c[1]) - eps <= b[1] <= max(a[1], c[1]) + eps
            ):
                changed = True
                continue
            keep.append(b)
        poly = keep

    if len(poly) < 3:
        return list(polygon)
    return poly


def visible_target_candidates(
    poly: Polygon,
    anchors: List[int],
    right_chain: List[int],
    right_order: Dict[int, int],
    params: PartitionParams,
) -> Dict[int, List[Tuple[int, float, int]]]:
    out: Dict[int, List[Tuple[int, float, int]]] = {}

    for ip in anchors:
        p = poly[ip]
        lst: List[Tuple[int, float, int]] = []
        for iq in right_chain:
            q = poly[iq]
            if q[0] <= p[0] + 1e-12:
                continue
            if not segment_is_visible_inside_polygon(poly, p, q, mid_samples=params.visibility_mid_samples):
                continue
            dy = abs(q[1] - p[1])
            dx = abs(q[0] - p[0])
            order = right_order.get(iq, -1)
            if order < 0:
                continue
            cost = dy + 0.02 * dx
            lst.append((iq, cost, order))

        lst.sort(key=lambda t: (t[1], t[2]))
        out[ip] = lst[:max(8, int(params.visibility_max_candidates))]

    return out


def _normalize_circumscribe_mode(mode: str) -> str:
    m = str(mode).strip().lower()
    if m in ("general", "gen", "generic"):
        return "general"
    if m in ("symmetric", "sym", "duicheng", "对称"):
        return "symmetric"
    raise ValueError("circumscribe_mode must be 'symmetric' or 'general'.")


def concave_decomp_pea(
    scale_ratio_collision: float = 1.025,
    num_vertices: int = 24,
    plot: bool = True,
    circumscribe_mode: str = "symmetric",
    general_num_points: int = 240,
) -> None:
    plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False

    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    input_xlsx = data_dir / "outer_polygon_vertices_pea.xlsx"
    output_xlsx = data_dir / "convex_partition_pea.xlsx"
    mode = _normalize_circumscribe_mode(circumscribe_mode)
    if mode == "general":
        from shape_decomposition.circum_concave_general import circum_concave_general
        circum_concave_general(
            output_xlsx=input_xlsx,
            scale_ratio_collision=scale_ratio_collision,
            shape_name="pea",
            num_vertices=num_vertices,
            num_points=general_num_points,
            plot=plot,
        )
    else:
        from shape_decomposition.circum_pea import circum_pea
        circum_pea(
            output_xlsx=input_xlsx,
            scale_ratio_collision=scale_ratio_collision,
            num_vertices=num_vertices,
            plot=plot,
        )

    params = PartitionParams(
        fan_count=19,
        merge_enable=True,
        merge_horiz_dy_eps=0.40,
        merge_max_iter=80,
        plot=bool(plot),
        sheet_name="outer_polygon_pea",
        simplify_collinear_eps=1e-10,
        visibility_mid_samples=3,
        visibility_max_candidates=32,
    )

    poly = load_polygon_vertices_from_excel(input_xlsx, sheet_name=params.sheet_name)
    poly = PolygonUtils.ensure_ccw(poly)

    if not PolygonUtils.is_simple_polygon(poly):
        raise ValueError("Input polygon is not simple (self-intersecting).")

    parts = convex_partition_pea_vertex_fan_chain_dp(poly, params)
    parts = merge_horizontal_symmetric(parts, params)

    if params.plot:
        plot_partition(poly, parts)

    save_partition_to_excel(
        output_xlsx=output_xlsx,
        original=poly,
        parts=parts,
        params=params,
        source_xlsx=input_xlsx,
    )


def main() -> None:
    concave_decomp_pea()


if __name__ == "__main__":
    main()
