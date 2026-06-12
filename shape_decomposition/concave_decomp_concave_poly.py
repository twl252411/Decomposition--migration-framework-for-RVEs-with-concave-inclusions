# file: convex_partition_concave_poly.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import math
from pathlib import Path
import time
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from scipy.spatial import ConvexHull

PRECISION = 6
EPS = 1e-8

Point = Tuple[float, float]
Polygon = List[Point]


# =========================
# Geometry / Polygon utils
# =========================
class GeometryUtils:
    @staticmethod
    def calculate_signed_area(poly: Polygon) -> float:
        area = 0.0
        n = len(poly)
        for i in range(n):
            j = (i + 1) % n
            area += poly[i][0] * poly[j][1] - poly[j][0] * poly[i][1]
        return area

    @staticmethod
    def cross(o: Point, a: Point, b: Point) -> float:
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    @staticmethod
    def distance(a: Point, b: Point) -> float:
        return math.hypot(a[0] - b[0], a[1] - b[1])

    @staticmethod
    def round_pt(p: Point) -> Point:
        return (round(float(p[0]), PRECISION), round(float(p[1]), PRECISION))

    @staticmethod
    def mean_center(poly: Polygon) -> Point:
        x = sum(p[0] for p in poly) / max(1, len(poly))
        y = sum(p[1] for p in poly) / max(1, len(poly))
        return (x, y)

    @staticmethod
    def point_in_triangle(p: Point, tri: Polygon) -> bool:
        # barycentric, boundary-inclusive
        (ax, ay), (bx, by), (cx, cy) = tri
        px, py = p
        v0x, v0y = cx - ax, cy - ay
        v1x, v1y = bx - ax, by - ay
        v2x, v2y = px - ax, py - ay

        dot00 = v0x * v0x + v0y * v0y
        dot01 = v0x * v1x + v0y * v1y
        dot02 = v0x * v2x + v0y * v2y
        dot11 = v1x * v1x + v1y * v1y
        dot12 = v1x * v2x + v1y * v2y

        denom = dot00 * dot11 - dot01 * dot01
        if abs(denom) < 1e-20:
            return False

        inv = 1.0 / denom
        u = (dot11 * dot02 - dot01 * dot12) * inv
        v = (dot00 * dot12 - dot01 * dot02) * inv
        return (u >= -EPS) and (v >= -EPS) and (u + v <= 1.0 + EPS)


class PolygonUtils:
    @staticmethod
    def ensure_ccw(poly: Polygon, eps: float = 1e-12) -> Polygon:
        area = GeometryUtils.calculate_signed_area(poly)
        if abs(area) < eps:
            raise ValueError(f"Degenerate polygon (signed area ~ 0): {area:.3e}")
        return poly if area > 0 else list(reversed(poly))

    @staticmethod
    def is_convex(poly: Polygon) -> bool:
        poly = PolygonUtils.ensure_ccw(poly)
        n = len(poly)
        if n < 3:
            return False
        for i in range(n):
            a = poly[i]
            b = poly[(i + 1) % n]
            c = poly[(i + 2) % n]
            if GeometryUtils.cross(a, b, c) < -EPS:
                return False
        return True

    @staticmethod
    def is_point_inside_polygon(p: Point, poly: Polygon) -> bool:
        # ray casting + boundary inclusive
        x, y = p
        inside = False
        n = len(poly)
        for i in range(n):
            j = (i + 1) % n
            xi, yi = poly[i]
            xj, yj = poly[j]

            # boundary check
            if abs(GeometryUtils.cross(poly[i], poly[j], p)) <= 1e-10:
                if (min(xi, xj) - EPS <= x <= max(xi, xj) + EPS and
                        min(yi, yj) - EPS <= y <= max(yi, yj) + EPS):
                    return True

            if (yi > y) != (yj > y):
                x_int = (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi
                if x < x_int:
                    inside = not inside
        return inside

    @staticmethod
    def is_simple_polygon(poly: Polygon, timeout: float = 10.0) -> bool:
        start = time.time()
        n = len(poly)
        if n < 3:
            return False

        # duplicate vertices
        seen = set()
        for v in poly:
            rv = GeometryUtils.round_pt(v)
            if rv in seen:
                return False
            seen.add(rv)

        for i in range(n):
            if time.time() - start > timeout:
                return False
            a, b = poly[i], poly[(i + 1) % n]
            for j in range(i + 2, n):
                if time.time() - start > timeout:
                    return False
                c, d = poly[j], poly[(j + 1) % n]
                if (j + 1) % n == i:
                    continue
                if SegmentUtils.segment_intersect(a, b, c, d):
                    return False
        return True


class SegmentUtils:
    @staticmethod
    def segment_intersect(a: Point, b: Point, c: Point, d: Point, eps: float = EPS) -> bool:
        def ccw(A: Point, B: Point, C: Point) -> float:
            v = (B[0] - A[0]) * (C[1] - A[1]) - (B[1] - A[1]) * (C[0] - A[0])
            return 0.0 if abs(v) <= eps else v

        def on_segment(p: Point, q: Point, r: Point) -> bool:
            return (min(p[0], r[0]) - eps <= q[0] <= max(p[0], r[0]) + eps and
                    min(p[1], r[1]) - eps <= q[1] <= max(p[1], r[1]) + eps and
                    ccw(p, q, r) == 0.0)

        ab_c = ccw(a, b, c)
        ab_d = ccw(a, b, d)
        cd_a = ccw(c, d, a)
        cd_b = ccw(c, d, b)

        if (ab_c * ab_d < 0.0) and (cd_a * cd_b < 0.0):
            return True

        return (on_segment(a, c, b) or on_segment(a, d, b) or
                on_segment(c, a, d) or on_segment(c, b, d))

    @staticmethod
    def is_valid_diagonal(poly: Polygon, i: int, j: int) -> bool:
        n = len(poly)
        if abs(i - j) == 1 or (i == 0 and j == n - 1) or (j == 0 and i == n - 1):
            return False

        a, b = poly[i], poly[j]
        for k in range(n):
            c = poly[k]
            d = poly[(k + 1) % n]
            if k in (i, j) or (k + 1) % n in (i, j):
                continue
            if SegmentUtils.segment_intersect(a, b, c, d):
                return False

        mid = ((a[0] + b[0]) * 0.5, (a[1] + b[1]) * 0.5)
        return PolygonUtils.is_point_inside_polygon(mid, poly)


# =========================
# Triangulation (divide & conquer + ear clipping fallback)
# =========================
class TriangulationUtils:
    @staticmethod
    def divide_and_conquer_triangulation(poly: Polygon) -> List[Polygon]:
        poly = PolygonUtils.ensure_ccw(poly)
        if len(poly) < 3:
            return []
        if not PolygonUtils.is_simple_polygon(poly):
            raise ValueError("Non-simple polygon cannot be triangulated.")

        if len(poly) == 3:
            return [poly]

        n = len(poly)
        for i in range(n):
            j = (i + 2) % n
            if SegmentUtils.is_valid_diagonal(poly, i, j):
                sub1 = poly[i:j + 1] if i < j else poly[i:] + poly[:j + 1]
                sub2 = poly[j:i + 1] if j < i else poly[j:] + poly[:i + 1]
                return (TriangulationUtils.divide_and_conquer_triangulation(PolygonUtils.ensure_ccw(sub1)) +
                        TriangulationUtils.divide_and_conquer_triangulation(PolygonUtils.ensure_ccw(sub2)))

        return TriangulationUtils.optimized_ear_clipping(poly)

    @staticmethod
    def optimized_ear_clipping(poly: Polygon) -> List[Polygon]:
        poly = PolygonUtils.ensure_ccw(poly)
        n = len(poly)
        if n < 3:
            return []

        prev = [(i - 1) % n for i in range(n)]
        nxt = [(i + 1) % n for i in range(n)]
        alive = [True] * n
        remaining = n
        cur = 0
        tris: List[Polygon] = []

        def is_ear(i: int) -> bool:
            a = prev[i]
            b = i
            c = nxt[i]
            pa, pb, pc = poly[a], poly[b], poly[c]

            if GeometryUtils.cross(pa, pb, pc) <= EPS:
                return False

            tri = [pa, pb, pc]
            for k in range(n):
                if not alive[k] or k in (a, b, c):
                    continue
                if GeometryUtils.point_in_triangle(poly[k], tri):
                    return False
            return True

        guard = 0
        while remaining > 3 and guard < 200000:
            guard += 1
            if not alive[cur]:
                cur = nxt[cur]
                continue

            if is_ear(cur):
                a = prev[cur]
                b = cur
                c = nxt[cur]
                tris.append([poly[a], poly[b], poly[c]])

                # remove vertex b
                alive[b] = False
                nxt[a] = c
                prev[c] = a
                remaining -= 1
                cur = a
            else:
                cur = nxt[cur]

        # final triangle
        ids = [i for i in range(n) if alive[i]]
        if len(ids) == 3:
            tris.append([poly[ids[0]], poly[ids[1]], poly[ids[2]]])

        return tris


# =========================
# Convex merge (shared-edge greedy)
# =========================
class ConvexMerger:
    def __init__(self, triangles: List[Polygon], original_poly: Polygon):
        self.original_poly = PolygonUtils.ensure_ccw(original_poly)
        self.polygons = [PolygonUtils.ensure_ccw(t) for t in triangles]
        self.edge_map: Dict[Tuple[Point, Point], set] = defaultdict(set)
        self._rebuild_edge_map()

    def _rebuild_edge_map(self) -> None:
        self.edge_map.clear()
        for pid, poly in enumerate(self.polygons):
            m = len(poly)
            for i in range(m):
                a = GeometryUtils.round_pt(poly[i])
                b = GeometryUtils.round_pt(poly[(i + 1) % m])
                edge = tuple(sorted([a, b]))
                self.edge_map[edge].add(pid)

    def _shared_edges_sorted(self) -> List[Tuple[float, Tuple[Point, Point]]]:
        out = []
        for edge, ids in self.edge_map.items():
            if len(ids) >= 2:
                out.append((GeometryUtils.distance(edge[0], edge[1]), edge))
        out.sort(key=lambda x: -x[0])  # longer first
        return out

    @staticmethod
    def merge_two_convex(poly1: Polygon, poly2: Polygon) -> Optional[Polygon]:
        eps = 10 ** (-PRECISION)

        def same(u: Point, v: Point) -> bool:
            return abs(u[0] - v[0]) < eps and abs(u[1] - v[1]) < eps

        p1 = PolygonUtils.ensure_ccw(poly1)
        p2 = PolygonUtils.ensure_ccw(poly2)

        shared = None  # (i_in_p1, j_in_p2, mode)
        n1, n2 = len(p1), len(p2)
        for i in range(n1):
            a1, b1 = p1[i], p1[(i + 1) % n1]
            for j in range(n2):
                a2, b2 = p2[j], p2[(j + 1) % n2]
                if same(a1, b2) and same(b1, a2):
                    shared = (i, j, "rev")
                    break
                if same(a1, a2) and same(b1, b2):
                    shared = (i, j, "same")
                    break
            if shared:
                break

        if not shared:
            return None

        i, j, mode = shared
        # indices of a/b on p2
        if mode == "rev":
            ja = (j + 1) % n2  # a
            jb = j             # b
        else:
            ja = j             # a
            jb = (j + 1) % n2  # b

        merged: List[Point] = []

        # p1: walk from b to a (inclusive)
        k = (i + 1) % n1
        merged.append(p1[k])
        while k != i:
            k = (k + 1) % n1
            merged.append(p1[k])

        # p2: walk from a to b (exclusive endpoints)
        k = (ja + 1) % n2
        while k != jb:
            merged.append(p2[k])
            k = (k + 1) % n2

        # remove consecutive duplicates
        cleaned: List[Point] = []
        for pt in merged:
            if not cleaned:
                cleaned.append(pt)
            else:
                if abs(cleaned[-1][0] - pt[0]) > eps or abs(cleaned[-1][1] - pt[1]) > eps:
                    cleaned.append(pt)

        if len(cleaned) >= 2 and same(cleaned[0], cleaned[-1]):
            cleaned.pop()

        if len(cleaned) < 3:
            return None

        cleaned = PolygonUtils.ensure_ccw(cleaned)
        if not PolygonUtils.is_convex(cleaned):
            return None

        # convex hull check (keep your original behavior)
        try:
            if len(ConvexHull(cleaned).vertices) != len(cleaned):
                return None
        except Exception:
            return None

        return cleaned

    def smart_merge(self) -> List[Polygon]:
        while True:
            edges = self._shared_edges_sorted()
            if not edges:
                break

            merged_any = False

            for _, edge in edges:
                poly_ids = list(self.edge_map.get(edge, set()))
                if len(poly_ids) < 2:
                    continue

                # try pairs, prefer larger area first
                areas = [(pid, abs(GeometryUtils.calculate_signed_area(self.polygons[pid]))) for pid in poly_ids]
                areas.sort(key=lambda x: -x[1])
                ids_sorted = [pid for pid, _ in areas]

                for ii in range(len(ids_sorted)):
                    for jj in range(ii + 1, len(ids_sorted)):
                        pida, pidb = ids_sorted[ii], ids_sorted[jj]
                        merged = ConvexMerger.merge_two_convex(self.polygons[pida], self.polygons[pidb])
                        if merged is None:
                            continue

                        # inside-original check: vertices + midpoints (与你原先等价的稳健检查)
                        ok = True
                        for v in merged:
                            if not PolygonUtils.is_point_inside_polygon(v, self.original_poly):
                                ok = False
                                break
                        if ok:
                            for t in range(len(merged)):
                                a = merged[t]
                                b = merged[(t + 1) % len(merged)]
                                mid = ((a[0] + b[0]) * 0.5, (a[1] + b[1]) * 0.5)
                                if not PolygonUtils.is_point_inside_polygon(mid, self.original_poly):
                                    ok = False
                                    break
                        if not ok:
                            continue

                        # apply merge
                        self.polygons = [p for k, p in enumerate(self.polygons) if k not in {pida, pidb}] + [merged]
                        self._rebuild_edge_map()
                        merged_any = True
                        break
                    if merged_any:
                        break
                if merged_any:
                    break

            if not merged_any:
                break

        return self.polygons


# =========================
# IO / Visualization / Pipeline
# =========================
def convex_partition(poly: Polygon) -> List[Polygon]:
    triangles = TriangulationUtils.divide_and_conquer_triangulation(poly)
    if not triangles:
        return [poly]
    return ConvexMerger(triangles, poly).smart_merge()


def load_polygon_vertices_from_excel(xlsx_path: str, sheet_name: str) -> Polygon:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    verts: Polygon = []
    for row in rows[1:]:
        if row is None or len(row) < 2 or row[0] is None or row[1] is None:
            break
        verts.append((float(row[0]), float(row[1])))

    if len(verts) < 3:
        raise ValueError("Polygon vertex count < 3 in Excel.")
    return verts


def _copy_sheet(source_ws, target_wb: Workbook, title: str) -> None:
    target_ws = target_wb.create_sheet(title=title)
    for row in source_ws.iter_rows(values_only=True):
        target_ws.append(list(row))


def save_partition_to_excel(output_xlsx: str, sheet_name: str, poly_scale: Polygon, parts: List[Polygon],
                            input_xlsx: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(["Original Polygon X", "Original Polygon Y"])
    for x, y in poly_scale:
        ws.append([float(x), float(y)])

    for i, part in enumerate(parts, start=1):
        ws.append([])
        ws.append([f"Convex Part {i} X", f"Convex Part {i} Y"])
        for x, y in part:
            ws.append([float(x), float(y)])

    source_wb = load_workbook(input_xlsx, data_only=True)
    for sheet_title in ("curve_concave_poly", "area_concave_poly"):
        if sheet_title not in source_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_title}' not found in {input_xlsx}.")
        _copy_sheet(source_wb[sheet_title], wb, sheet_title)

    wb.save(output_xlsx)


def visualize_single_convex_partition(concave_polygon: Polygon, plot: bool = True) -> List[Polygon]:
    parts = convex_partition(concave_polygon)

    if plot:
        plt.figure(figsize=(10, 8))
        x, y = zip(*concave_polygon)
        plt.plot(x + (x[0],), y + (y[0],), "k--", lw=2, label="Original Polygon")

        colors = plt.cm.tab20.colors
        for i, part in enumerate(parts):
            px, py = zip(*part)
            plt.fill(px, py, color=colors[i % 20], alpha=0.3)
            plt.plot(px + (px[0],), py + (py[0],), "-", lw=2, color=colors[i % 20])
            plt.text(sum(p[0] for p in part) / len(part),
                     sum(p[1] for p in part) / len(part),
                     f"Part {i + 1}", ha="center", va="center", fontsize=10)

        plt.title(f"Convex Partition ({len(parts)} parts)", fontsize=14)
        plt.axis("equal")
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.show()

    return parts


# =========================
# main (keep last)
# =========================
def concave_decomp_concave_poly(scale_ratio_collision: float = 1.025, plot: bool = True) -> None:
    data_dir = Path(__file__).resolve().parents[1] / "intermediate_final_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    input_xlsx = data_dir / "outer_polygon_vertices_concave_poly.xlsx"
    input_sheet = "outer_polygon_concave_poly"
    output_xlsx = data_dir / "convex_partition_concave_poly.xlsx"
    from shape_decomposition.circum_concave_poly import circum_concave_poly
    circum_concave_poly(output_xlsx=input_xlsx, scale_ratio_collision=scale_ratio_collision)

    poly = load_polygon_vertices_from_excel(input_xlsx, input_sheet)
    poly = PolygonUtils.ensure_ccw(poly)
    if not PolygonUtils.is_simple_polygon(poly):
        raise ValueError("多边形存在自交或重复点，请先修正 Excel 顶点顺序/数据。")

    parts = visualize_single_convex_partition(poly, plot=plot)
    save_partition_to_excel(output_xlsx, input_sheet, poly, parts, str(input_xlsx))


def main() -> None:
    concave_decomp_concave_poly()


if __name__ == "__main__":
    main()
