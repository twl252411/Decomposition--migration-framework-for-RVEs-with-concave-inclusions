from typing import List, Optional, Tuple
import numpy as np
import csv
from openpyxl import load_workbook, Workbook
from pathlib import Path
import time
from .geom2d import aabb_from_points, convex_hull_center_radius
from .periodic import periodic_shift_candidates, polygon_vertices_world
from .rve_types import ShapeTemplate, PolygonShape, is_polygon_convex, local_normals_from_polygon
from shape_decomposition.circum_lobular2 import calculate_polygon_area

def load_template_from_excel(xlsx_path: str, sheet_name: str) -> ShapeTemplate:
    wb = load_workbook(xlsx_path, data_only=True)
    sh = wb[sheet_name]
    original, parts = parse_polygon_sheet(sh)

    lc, r = convex_hull_center_radius(original)
    aabb_whole = aabb_from_points(original)

    part_centers = np.zeros((len(parts), 2), dtype=np.float64)
    part_radii = np.zeros((len(parts),), dtype=np.float64)
    aabb_parts = np.zeros((len(parts), 4), dtype=np.float64)
    part_normals = []

    for i, p in enumerate(parts):
        c, rr = convex_hull_center_radius(p)
        part_centers[i] = c
        part_radii[i] = rr
        aabb_parts[i] = aabb_from_points(p)
        part_normals.append(local_normals_from_polygon(np.asarray(p, dtype=np.float64)))

    return ShapeTemplate(
        aabb_parts=aabb_parts,
        aabb_whole=aabb_whole,
        local_center=lc,
        part_centers=part_centers,
        part_normals=part_normals,
        part_radii=part_radii,
        parts=[np.asarray(p, dtype=np.float64) for p in parts],
        radius=float(r),
        vertices=np.asarray(original, dtype=np.float64),
        is_convex=is_polygon_convex(np.asarray(original, dtype=np.float64)),
    )


def parse_polygon_sheet(sheet) -> Tuple[np.ndarray, List[np.ndarray]]:
    rows = list(sheet.iter_rows(values_only=True))
    return parse_polygon_rows(rows)


def parse_polygon_rows(rows) -> Tuple[np.ndarray, List[np.ndarray]]:
    original = []
    i = 1
    while i < len(rows):
        r = rows[i]
        if r is None or len(r) < 2 or r[0] is None or r[1] is None:
            break
        original.append((float(r[0]), float(r[1])))
        i += 1
    if len(original) < 3:
        raise ValueError("Original polygon has too few vertices.")
    original = np.asarray(original, dtype=np.float64)

    parts: List[np.ndarray] = []
    cur: List[Tuple[float, float]] = []
    for r in rows[i + 1 :]:
        if r is None:
            continue
        c0 = r[0] if len(r) > 0 else None
        c1 = r[1] if len(r) > 1 else None
        if c0 is None and c1 is None:
            continue
        if c0 is not None and "Convex Part" in str(c0):
            if len(cur) >= 3:
                parts.append(np.asarray(cur, dtype=np.float64))
            cur = []
            continue
        if c0 is None or c1 is None:
            continue
        cur.append((float(c0), float(c1)))
    if len(cur) >= 3:
        parts.append(np.asarray(cur, dtype=np.float64))
    if len(parts) == 0:
        raise ValueError("No convex parts found in sheet.")
    return original, parts


def load_template_from_csv(csv_path: str) -> ShapeTemplate:
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = [row for row in reader]
    original, parts = parse_polygon_rows(rows)

    lc, r = convex_hull_center_radius(original)
    aabb_whole = aabb_from_points(original)

    part_centers = np.zeros((len(parts), 2), dtype=np.float64)
    part_radii = np.zeros((len(parts),), dtype=np.float64)
    aabb_parts = np.zeros((len(parts), 4), dtype=np.float64)
    part_normals = []

    for i, p in enumerate(parts):
        c, rr = convex_hull_center_radius(p)
        part_centers[i] = c
        part_radii[i] = rr
        aabb_parts[i] = aabb_from_points(p)
        part_normals.append(local_normals_from_polygon(np.asarray(p, dtype=np.float64)))

    return ShapeTemplate(
        aabb_parts=aabb_parts,
        aabb_whole=aabb_whole,
        local_center=lc,
        part_centers=part_centers,
        part_normals=part_normals,
        part_radii=part_radii,
        parts=[np.asarray(p, dtype=np.float64) for p in parts],
        radius=float(r),
        vertices=np.asarray(original, dtype=np.float64),
        is_convex=is_polygon_convex(np.asarray(original, dtype=np.float64)),
    )


def _write_polygon_sheet(ws, shapes: List[PolygonShape]) -> None:
    ws.append(["polygon_id", "tx", "ty", "angle_rad"])
    for i, s in enumerate(shapes):
        ws.append([i + 1, float(s.translation[0]), float(s.translation[1]), float(s.angle)])


def _write_polygon_periodic_sheet(ws, shapes: List[PolygonShape], rve_size: float) -> None:
    ws.append(["polygon_id", "image_id", "shift_x", "shift_y", "tx", "ty", "angle_rad"])
    for i, s in enumerate(shapes):
        base_vertices = polygon_vertices_world(s)
        shifts = periodic_shift_candidates(base_vertices, rve_size)
        for image_id, (dx, dy) in enumerate(shifts, start=1):
            ws.append(
                [
                    i + 1,
                    image_id,
                    float(dx),
                    float(dy),
                    float(s.translation[0] + dx),
                    float(s.translation[1] + dy),
                    float(s.angle),
                ]
            )


def save_polygon_data(
    shapes: List[PolygonShape],
    filename: str,
    sheet_title: str,
    *,
    include_periodic_images: bool = False,
    rve_size: Optional[float] = None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_polygon_sheet(ws, shapes)

    if include_periodic_images:
        if rve_size is None:
            raise ValueError("rve_size must be provided when include_periodic_images=True.")
        ws_periodic = wb.create_sheet(title=f"{sheet_title}_periodic")
        _write_polygon_periodic_sheet(ws_periodic, shapes, float(rve_size))
    wb.save(filename)


def save_polygon_data_txt(
    shapes: List[PolygonShape],
    filename: str,
    *,
    include_periodic_images: bool = False,
    rve_size: Optional[float] = None,
) -> None:
    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        for i, s in enumerate(shapes):
            writer.writerow([float(s.translation[0]), float(s.translation[1]), float(s.angle)])

        if include_periodic_images:
            if rve_size is None:
                raise ValueError("rve_size must be provided when include_periodic_images=True.")
            for i, s in enumerate(shapes):
                base_vertices = polygon_vertices_world(s)
                shifts = periodic_shift_candidates(base_vertices, float(rve_size))
                for dx, dy in shifts:
                    if dx == 0.0 and dy == 0.0:
                        continue
                    writer.writerow(
                        [
                            float(s.translation[0] + dx),
                            float(s.translation[1] + dy),
                            float(s.angle),
                        ]
                    )


def load_polygon_vertices_from_excel(xlsx_path: str, sheet_name: str) -> np.ndarray:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    vertices = []
    for row in rows[1:]:
        if row is None or len(row) < 2 or row[0] is None or row[1] is None:
            break
        vertices.append((float(row[0]), float(row[1])))
    if len(vertices) < 3:
        raise ValueError("Polygon vertex count < 3 in Excel.")
    return np.asarray(vertices, dtype=np.float64)


def load_area_values_from_excel(xlsx_path: str, sheet_name: str) -> dict[str, float]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Area sheet does not contain header and value rows.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    values = rows[1]
    area_values: dict[str, float] = {}
    for header, value in zip(headers, values):
        if not header or value is None:
            continue
        area_values[header] = float(value)
    if not area_values:
        raise ValueError("Area sheet contains no values.")
    return area_values


def save_run_metrics_txt(
    filename: Path,
    sections: List[Tuple[str, List[Tuple[str, object]]]],
    *,
    start_time: Optional[float] = None,
) -> None:
    def format_value(key: str, value: object) -> str:
        if value is None:
            return "N/A"
        if isinstance(value, bool):
            return "True" if value else "False"
        if isinstance(value, (int, np.integer)):
            return str(int(value))
        if isinstance(value, (float, np.floating)):
            if key.startswith("t_"):
                return f"{float(value):.6f} s"
            return f"{float(value):.6g}"
        return str(value)

    with open(filename, "w", encoding="utf-8", newline="") as f:
        for section, items in sections:
            if section:
                f.write(f"[{section}]\n")
            for key, value in items:
                if key == "t_tot" and value is None and start_time is not None:
                    value = time.time() - start_time
                f.write(f"{key}\t{format_value(key, value)}\n")
            f.write("\n")


def load_curve_outline_and_area(demo_shape: str, template: ShapeTemplate,
                                 data_dir: Path) -> tuple[np.ndarray, float]:
    if demo_shape in ("pea", "lobular2", "concave_poly"):
        curve_path = data_dir / f"convex_partition_{demo_shape}.xlsx"
        curve_sheet = f"curve_{demo_shape}"
        area_sheet = f"area_{demo_shape}"
        try:
            outline_vertices = load_polygon_vertices_from_excel(curve_path, curve_sheet)
            area_values = load_area_values_from_excel(curve_path, area_sheet)
            area_curve = area_values.get("area_curve")
            return outline_vertices, float(area_curve)
        except (FileNotFoundError, KeyError, ValueError) as exc:
            print(f"Warning: failed to load curve data for {demo_shape}: {exc}")
            area_poly = calculate_polygon_area(template.vertices)
            return template.vertices, float(area_poly)

    raise ValueError("demo_shape must be 'pea', 'lobular2', or 'concave_poly'")
